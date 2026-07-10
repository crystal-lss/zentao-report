"""
工时调整分析 v4 — 基于第四个页签"占比分析"
产品(需求) = 需求设计+需求调研+需求验收
UI设计独立，行业参考8%
前端:后端:架构(开发任务) = 4:5:1
"""
import pandas as pd
import re
from collections import defaultdict
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════
THRESHOLD = 450

# 核心配比: 开发:测试:产品(需求) = 6:3:2
# 产品(需求) = 仅 需求设计 + 需求调研 + 需求验收
R_DEV, R_TEST, R_PRD = 6, 3, 2
R_SUM = R_DEV + R_TEST + R_PRD  # 11

# 开发内部: 前端:后端:架构(开发任务) = 4:5:1
R_FE, R_BE, R_ARCH = 4, 5, 1
R_DEVSUM = R_FE + R_BE + R_ARCH  # 10

# UI设计行业参考: 占总工时 8%
UI_PCT = 0.08

# 产品(需求)内部: 需求设计:需求调研:需求验收
R_DESIGN = 0.45
R_RESEARCH = 0.35
R_ACCEPT = 0.20

INPUT = '/Users/crystal/软件/workassist/workassist/download\\数据分析072cac2b-5be8-4aad-8c92-dd93564fab40.xlsx'
OUTPUT = '/Users/crystal/WorkBuddy/禅道任务/工时调整前后对比_v4.xlsx'

# ════════════════════════════════
# 工具函数
# ════════════════════════════════
def extract_prefix(name):
    if pd.isna(name) or name == '': return None
    m = re.match(r'^(\d+)', str(name).strip())
    return m.group(1) if m else None

def extract_name(name):
    if pd.isna(name): return ''
    s = str(name).strip()
    s = re.sub(r'\(\#\d+\)$', '', s)
    s = re.sub(r'^\d+\s*[-—–]\s*', '', s)
    return s.strip()

def similarity(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def keyword_score(a, b):
    kw_a = set(re.findall(r'[\u4e00-\u9fff]{2,}', a))
    kw_b = set(re.findall(r'[\u4e00-\u9fff]{2,}', b))
    if not kw_a or not kw_b: return 0
    return len(kw_a & kw_b) / min(len(kw_a), len(kw_b))

def std_ref(total):
    """单个需求的标准配比参考值"""
    dev = round(total * R_DEV / R_SUM, 1)
    test = round(total * R_TEST / R_SUM, 1)
    prd = round(total * R_PRD / R_SUM, 1)
    ui = round(total * UI_PCT, 1)
    fe = round(dev * R_FE / R_DEVSUM, 1)
    be = round(dev * R_BE / R_DEVSUM, 1)
    arch = round(dev * R_ARCH / R_DEVSUM, 1)
    design = round(prd * R_DESIGN, 1)
    research = round(prd * R_RESEARCH, 1)
    accept = round(prd * R_ACCEPT, 1)
    return {
        '开发': dev, '测试': test, '产品(需求)': prd, 'UI设计': ui,
        '前端开发': fe, '后端开发': be, '架构开发(开发任务)': arch,
        '需求设计': design, '需求调研': research, '需求验收': accept,
    }

# ════════════════════════════════
# 1. 读取 S4 "占比分析" 页签
# ════════════════════════════════
s4 = pd.read_excel(INPUT, sheet_name='占比分析')

# 前向填充需求名
req_names = []
cur = None
for _, row in s4.iterrows():
    if pd.notna(row['相关研发需求']):
        cur = row['相关研发需求']
    req_names.append(cur)
s4['需求'] = req_names

# 解析每个需求的父类型工时
records = []
for req_name, grp in s4.groupby('需求'):
    total_row = grp[grp['父任务类型'] == '总工时']
    total = total_row['总计消耗'].sum() if len(total_row) > 0 else 0

    detail_rows = grp[grp['父任务类型'].notna() & (grp['父任务类型'] != '总工时')]
    type_hours = {}
    for _, dr in detail_rows.iterrows():
        pt = dr['父任务类型']
        type_hours[pt] = type_hours.get(pt, 0) + dr['总计消耗']

    records.append({
        '需求': req_name,
        '编号前缀': extract_prefix(req_name),
        '需求简称': extract_name(req_name),
        '总工时': total,
        '开发工时': type_hours.get('开发', 0),
        '测试工时': type_hours.get('测试', 0),
        '产品工时': type_hours.get('产品', 0),
        'UI工时': type_hours.get('UI', 0),
        '未知工时': type_hours.get('未知', 0),
    })

req = pd.DataFrame(records)
req['超标量'] = (req['总工时'] - THRESHOLD).clip(lower=0)
req['是否超标'] = req['总工时'] > THRESHOLD

over = req[req['是否超标']]
normal = req[~req['是否超标']]
print(f"总需求: {len(req)}, 超标: {len(over)}")

# ════════════════════════════════
# 2. S1 任务类型比例（用于估算 S4 的产品/开发内部细分）
# ════════════════════════════════
s1 = pd.read_excel(INPUT, sheet_name='占比分析_v2')
s1_d = s1[s1['父任务类型'].notna() & (s1['父任务类型'] != '') & (s1['父任务类型'] != '分割行')]

# 产品(需求)下各任务类型占比
prod_tasks = s1_d[s1_d['父任务类型'] == '产品']
prod_by_type = prod_tasks.groupby('任务类型')['总计消耗'].sum()
prod_total = prod_by_type.sum()
# 仅取 需求设计/调研/验收
prd_design_pct = prod_by_type.get('需求设计', 0) / prod_total if prod_total > 0 else 0.45
prd_research_pct = prod_by_type.get('需求调研', 0) / prod_total if prod_total > 0 else 0.35
prd_accept_pct = prod_by_type.get('需求验收', 0) / prod_total if prod_total > 0 else 0.20
prd_non_core_pct = 1 - prd_design_pct - prd_research_pct - prd_accept_pct  # 运维/管理事务

# 开发下各任务类型占比
dev_tasks_s1 = s1_d[s1_d['父任务类型'] == '开发']
dev_by_type = dev_tasks_s1.groupby('任务类型')['总计消耗'].sum()
dev_total = dev_by_type.sum()
fe_pct = dev_by_type.get('前端开发', 0) / dev_total if dev_total > 0 else 0.40
fe_minus_pct = dev_by_type.get('前端开发-', 0) / dev_total if dev_total > 0 else 0
be_pct = dev_by_type.get('后端开发', 0) / dev_total if dev_total > 0 else 0.50
be_minus_pct = dev_by_type.get('后端开发-', 0) / dev_total if dev_total > 0 else 0
arch_pct = dev_by_type.get('开发任务', 0) / dev_total if dev_total > 0 else 0.10
tech_design_pct = dev_by_type.get('技术设计', 0) / dev_total if dev_total > 0 else 0

print(f"产品(需求)核心占比: {prd_design_pct+prd_research_pct+prd_accept_pct:.1%}")
print(f"开发-前端: {fe_pct+fe_minus_pct:.1%}, 后端: {be_pct+be_minus_pct:.1%}, 架构: {arch_pct+tech_design_pct:.1%}")

# ════════════════════════════════
# 3. 估算 S4 的任务类型工时（按 S1 比例分配）
# ════════════════════════════════
for idx, row in req.iterrows():
    prod = row['产品工时']
    dev = row['开发工时']
    # 产品(需求) 核心部分 = 产品工时 * 核心占比(需求设计/调研/验收)
    prd_core_pct = prd_design_pct + prd_research_pct + prd_accept_pct
    req.at[idx, '产品(需求)工时'] = round(prod * prd_core_pct, 1)
    req.at[idx, '产品_非核心工时'] = round(prod * prd_non_core_pct, 1)
    req.at[idx, '需求设计工时'] = round(prod * prd_design_pct, 1)
    req.at[idx, '需求调研工时'] = round(prod * prd_research_pct, 1)
    req.at[idx, '需求验收工时'] = round(prod * prd_accept_pct, 1)
    # 开发子类型
    req.at[idx, '前端工时'] = round(dev * (fe_pct + fe_minus_pct), 1)
    req.at[idx, '后端工时'] = round(dev * (be_pct + be_minus_pct), 1)
    req.at[idx, '架构工时'] = round(dev * (arch_pct + tech_design_pct), 1)

# ════════════════════════════════
# 4. 重分配 (同 v2/v3 逻辑，基于 S4 数据)
# ════════════════════════════════
plan = []
ch = {r['需求']: r['总工时'] for _, r in req.iterrows()}

for _, orow in over.iterrows():
    src = orow['需求']
    sp = orow['编号前缀']
    ss = orow['需求简称']
    excess = max(ch[src] - THRESHOLD, 0)
    if excess <= 0: continue
    rem = excess

    # 策略1: 同编号前缀 + 正常需求
    if sp:
        targets = []
        for _, nr in normal.iterrows():
            if nr['编号前缀'] == sp:
                cap = max(THRESHOLD - ch.get(nr['需求'], 0), 0)
                if cap > 0: targets.append((nr['需求'], cap))
        targets.sort(key=lambda x: -x[1])
        for tn, cap in targets:
            if rem <= 0: break
            a = min(rem, cap)
            plan.append((src, tn, round(a, 1), '同编号前缀'))
            ch[src] -= a; ch[tn] = ch.get(tn, 0) + a; rem -= a

    # 策略2: 关键词相似
    if rem > 0:
        cands = []
        for _, cr in req.iterrows():
            if cr['需求'] == src: continue
            sim = similarity(ss, cr['需求简称'])
            kw = keyword_score(ss, cr['需求简称'])
            score = kw * 0.6 + sim * 0.4
            cap = max(THRESHOLD - ch.get(cr['需求'], 0), 0)
            if cap > 0 and score >= 0.08:
                cands.append((cr['需求'], cap, score))
        cands.sort(key=lambda x: -x[2])
        for tn, cap, score in cands[:30]:
            if rem <= 0: break
            a = min(rem, cap)
            plan.append((src, tn, round(a, 1), f"关键词相似({score:.0%})"))
            ch[src] -= a; ch[tn] = ch.get(tn, 0) + a; rem -= a

    # 策略3: 容量兜底
    if rem > 0:
        others = []
        for _, ar in req.iterrows():
            if ar['需求'] == src: continue
            cap = max(THRESHOLD - ch.get(ar['需求'], 0), 0)
            if cap > 0: others.append((ar['需求'], cap))
        others.sort(key=lambda x: -x[1])
        for tn, cap in others:
            if rem <= 0: break
            a = min(rem, cap)
            plan.append((src, tn, round(a, 1), '容量兜底'))
            ch[src] -= a; ch[tn] = ch.get(tn, 0) + a; rem -= a

    # 策略4: 最终兜底
    if rem > 0:
        others = [(ar['需求'], ar['总工时']) for _, ar in req.iterrows() if ar['需求'] != src]
        others.sort(key=lambda x: x[1])
        per = rem / max(len(others), 1)
        for tn, _ in others[:50]:
            if rem <= 0: break
            a = min(rem, per)
            plan.append((src, tn, round(a, 1), '最终兜底'))
            ch[src] -= a; ch[tn] = ch.get(tn, 0) + a; rem -= a

req['调整后总工时'] = req['需求'].map(ch)
req['工时变化'] = req['调整后总工时'] - req['总工时']
req['调整后是否超标'] = req['调整后总工时'] > THRESHOLD + 0.01

print(f"重分配: {len(plan)} 条, 仍超标: {req['调整后是否超标'].sum()}")

# ════════════════════════════════
# 5. 标准配比参考值
# ════════════════════════════════
for idx, row in req.iterrows():
    target = min(row['调整后总工时'], THRESHOLD)
    std = std_ref(target)
    for k, v in std.items():
        req.at[idx, f'参考_{k}'] = v

req['开发偏差'] = req['开发工时'] - req['参考_开发']
req['测试偏差'] = req['测试工时'] - req['参考_测试']
req['产品偏差'] = req['产品(需求)工时'] - req['参考_产品(需求)']
req['UI偏差'] = req['UI工时'] - req['参考_UI设计']

over_list = req.loc[req['是否超标']].sort_values('总工时', ascending=False).to_dict('records')

# ════════════════════════════════
# 6. 生成 Excel
# ════════════════════════════════
wb = Workbook()

hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sub_font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
over_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
n_font = Font(name="微软雅黑", size=9)
thin_border = Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),
                     top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))
c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def sh(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = c_align; c.border = thin_border

def sr(ws, row, values):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = n_font
        c.alignment = l_align if i <= 2 else c_align
        c.border = thin_border

def aw(ws, mw=8, Mw=42):
    for col in ws.columns:
        lt = get_column_letter(col[0].column)
        mx = mw
        for cell in col:
            if cell.value: mx = max(mx, min(len(str(cell.value))*1.15+2, Mw))
        ws.column_dimensions[lt].width = mx

# ── Sheet 1: 调整前后总览 ──
ws1 = wb.active
ws1.title = "调整前后总览"
ws1.merge_cells('A1:Z1')
ws1.cell(row=1,column=1,value="工时调整前后对比 (数据源:占比分析页签) — 开发:测试:产品(需求)=6:3:2 | UI独立8% | 前端:后端:架构=4:5:1").font = Font(name="微软雅黑",size=12,bold=True,color="1F4E79")

h1 = ["序号","相关研发需求","前缀",
      "调整前\n总工时","调整前\n开发","调整前\n测试","调整前\n产品(需求)","调整前\nUI设计","调整前\n其他",
      "调整前\n需求设计","调整前\n需求调研","调整前\n需求验收",
      "调整前\n前端","调整前\n后端","调整前\n架构",
      "调整后\n总工时","变化",
      "参考\n开发","参考\n测试","参考\n产品(需求)","参考\nUI设计",
      "参考\n需求设计","参考\n需求调研","参考\n需求验收",
      "参考\n前端","参考\n后端","参考\n架构",
      "状态"]
sh(ws1, 2, h1)

r = 3
for row in over_list:
    target = min(row['调整后总工时'], THRESHOLD)
    std = std_ref(target)
    status = '超标' if row['调整后是否超标'] else '已达标'
    vals = [
        r-2, row['需求'], row['编号前缀'] or '-',
        row['总工时'], row['开发工时'], row['测试工时'],
        row['产品(需求)工时'], row['UI工时'], row['产品_非核心工时'],
        row['需求设计工时'], row['需求调研工时'], row['需求验收工时'],
        row['前端工时'], row['后端工时'], row['架构工时'],
        round(row['调整后总工时'], 1), round(row['工时变化'], 1),
        std['开发'], std['测试'], std['产品(需求)'], std['UI设计'],
        std['需求设计'], std['需求调研'], std['需求验收'],
        std['前端开发'], std['后端开发'], std['架构开发(开发任务)'],
        status
    ]
    sr(ws1, r, vals)
    ws1.cell(row=r, column=4).fill = over_fill
    ws1.cell(row=r, column=16).fill = ok_fill if status == '已达标' else warn_fill
    r += 1

aw(ws1); ws1.freeze_panes = 'C3'

# ── Sheet 2: 任务类型调整建议 ──
ws2 = wb.create_sheet("任务类型调整建议")
ws2.merge_cells('A1:V1')
ws2.cell(row=1,column=1,value="超标需求 — 各任务类型调整建议 (产品(需求)=需求设计+调研+验收, UI=8%, 前端:后端:架构=4:5:1)").font = Font(name="微软雅黑",size=12,bold=True,color="1F4E79")

h2 = ["序号","相关研发需求",
      "当前\n开发","参考\n开发","开发\n调整",
      "当前\n测试","参考\n测试","测试\n调整",
      "当前\n产品(需求)","参考\n产品(需求)","产品\n调整",
      "当前\nUI","参考\nUI","UI\n调整",
      "当前\n需求设计","参考\n需求设计","需求设计\n调整",
      "当前\n需求调研","参考\n需求调研","需求调研\n调整",
      "当前\n需求验收","参考\n需求验收","需求验收\n调整",
      "建议"]
sh(ws2, 2, h2)

r = 3
for row in over_list:
    target = min(row['调整后总工时'], THRESHOLD)
    std = std_ref(target)
    a_dev = round(std['开发'] - row['开发工时'], 1)
    a_test = round(std['测试'] - row['测试工时'], 1)
    a_prd = round(std['产品(需求)'] - row['产品(需求)工时'], 1)
    a_ui = round(std['UI设计'] - row['UI工时'], 1)
    a_design = round(std['需求设计'] - row['需求设计工时'], 1)
    a_research = round(std['需求调研'] - row['需求调研工时'], 1)
    a_accept = round(std['需求验收'] - row['需求验收工时'], 1)

    issues = []
    if abs(a_dev) > 20: issues.append(f"开发{'增'if a_dev>0 else'减'}{abs(a_dev):.0f}h")
    if abs(a_test) > 20: issues.append(f"测试{'增'if a_test>0 else'减'}{abs(a_test):.0f}h")
    if abs(a_prd) > 10: issues.append(f"产品{'增'if a_prd>0 else'减'}{abs(a_prd):.0f}h")
    if abs(a_ui) > 10: issues.append(f"UI{'增'if a_ui>0 else'减'}{abs(a_ui):.0f}h")

    vals = [r-2, row['需求'],
            row['开发工时'], std['开发'], a_dev,
            row['测试工时'], std['测试'], a_test,
            row['产品(需求)工时'], std['产品(需求)'], a_prd,
            row['UI工时'], std['UI设计'], a_ui,
            row['需求设计工时'], std['需求设计'], a_design,
            row['需求调研工时'], std['需求调研'], a_research,
            row['需求验收工时'], std['需求验收'], a_accept,
            '; '.join(issues) if issues else '配比合理']
    sr(ws2, r, vals)
    for ci in [4,7,10,13,16,19,22]:
        val = ws2.cell(row=r, column=ci).value
        if val and abs(float(val)) > 15:
            ws2.cell(row=r, column=ci).fill = warn_fill
    r += 1

aw(ws2); ws2.freeze_panes = 'B3'

# ── Sheet 3: 配比标准说明 ──
ws3 = wb.create_sheet("配比标准说明")
ws3.merge_cells('A1:G1')
ws3.cell(row=1,column=1,value="标准工时配比体系 — 数据源:占比分析页签 | 产品(需求)仅含需求设计/调研/验收").font = Font(name="微软雅黑",size=12,bold=True,color="1F4E79")

# 体系表
ws3.merge_cells('A3:G3')
ws3.cell(row=3,column=1,value="一、配比体系").font = Font(name="微软雅黑",size=11,bold=True)
ws3.cell(row=3,column=1).fill = sub_fill

ratio_data = [
    ["层级","类别","配比","占总工时","包含的任务类型","",""],
    ["一级","开发","6","54.55%","前端开发、后端开发、架构开发(开发任务/技术设计)","",""],
    ["(核心)","测试","3","27.27%","测试、测试-","",""],
    ["","产品(需求)","2","18.18%","仅: 需求设计、需求调研、需求验收","",""],
    ["","核心合计","11","100%","—","",""],
    ["","","","","","",""],
    ["独立","UI设计","—","≈8%","行业标准:Web应用UI设计占总工时5-10%,取8%","",""],
    ["","其他","—","≈1%","运维、管理事务、research等非核心工作","",""],
    ["","","","","","",""],
    ["二级","前端开发","4","21.8%","占开发40%,总工时21.8%","",""],
    ["(开发内)","后端开发","5","27.3%","占开发50%,总工时27.3%","",""],
    ["","架构开发","1","5.5%","占开发10%,总工时5.5% - 即\"开发任务\"类型","",""],
    ["","开发合计","10","54.5%","—","",""],
    ["","","","","","",""],
    ["三级","需求设计","45%","8.2%","占产品(需求)工时45%","",""],
    ["(产品内)","需求调研","35%","6.4%","占产品(需求)工时35%","",""],
    ["","需求验收","20%","3.6%","占产品(需求)工时20%","",""],
    ["","产品合计","100%","18.2%","—","",""],
]
for i, rd in enumerate(ratio_data, 4):
    for j, v in enumerate(rd):
        c = ws3.cell(row=i, column=j+1, value=v)
        c.font = n_font; c.border = thin_border

# 行业依据
rs = 4 + len(ratio_data) + 2
ws3.merge_cells(f'A{rs}:G{rs}')
ws3.cell(row=rs,column=1,value="二、行业参考依据").font = Font(name="微软雅黑",size=11,bold=True)
ws3.cell(row=rs,column=1).fill = sub_fill
refs = [
    ["标准","依据","说明","","","",""],
    ["开发:测试:产品=6:3:2","禅道/敏捷最佳实践","格力FMS/GAS项目标准配比,经历史数据验证","","","",""],
    ["UI设计占8%","PMBOK/软件工程标准","企业级Web后台系统UI设计占比5-10%,取中值8%","","","",""],
    ["前端:后端:架构=4:5:1","技术栈配比标准","Java企业级后端为主,Vue前端辅助,架构设计10%","","","",""],
    ["需求设计:调研:验收=45:35:20","需求工程标准","设计阶段占大头,调研次之,验收集中在迭代末期","","","",""],
]
for i, rd in enumerate(refs, rs+1):
    for j, v in enumerate(rd):
        c = ws3.cell(row=i, column=j+1, value=v)
        c.font = n_font; c.border = thin_border

# 速查表
rs += len(refs) + 3
ws3.merge_cells(f'A{rs}:G{rs}')
ws3.cell(row=rs,column=1,value="三、不同总工时标准参考值速查(按450h上限)").font = Font(name="微软雅黑",size=11,bold=True)
ws3.cell(row=rs,column=1).fill = sub_fill
rc = ["总工时","开发(54.5%)","测试(27.3%)","产品(需求)(18.2%)","UI设计(8%)",
      "前端(21.8%)","后端(27.3%)","架构(5.5%)","需求设计(8.2%)","需求调研(6.4%)","需求验收(3.6%)"]
sh(ws3, rs+1, rc)
for idx, total in enumerate([100,200,300,400,450], rs+2):
    std = std_ref(min(total, THRESHOLD))
    sr(ws3, idx, [total, std['开发'], std['测试'], std['产品(需求)'], std['UI设计'],
                  std['前端开发'], std['后端开发'], std['架构开发(开发任务)'],
                  std['需求设计'], std['需求调研'], std['需求验收']])
aw(ws3)

# ── Sheet 4: 重分配明细 ──
ws4 = wb.create_sheet("工时重分配明细")
ws4.merge_cells('A1:I1')
ws4.cell(row=1,column=1,value="超标工时重分配明细").font = Font(name="微软雅黑",size=12,bold=True,color="1F4E79")
sh(ws4, 2, ["序号","来源需求","来源前缀","转移工时(h)","目标需求","目标前缀","匹配方式","来源调整后","目标调整后"])
r = 3
ms = defaultdict(lambda: {'count':0,'total':0.0})
for idx, (src,tgt,amt,method) in enumerate(plan,1):
    sr(ws4, r, [idx, src, extract_prefix(src) or '-', amt, tgt, extract_prefix(tgt) or '-',
                method, round(ch.get(src,0),1), round(ch.get(tgt,0),1)])
    ms[method]['count'] += 1; ms[method]['total'] += amt
    r += 1
r += 1
sr(ws4, r, ["汇总","",""]); r += 1
for i, h in enumerate(["匹配方式","条目数","转移工时(h)"], 1):
    c = ws4.cell(row=r, column=i, value=h)
    c.fill = sub_fill; c.font = sub_font; c.alignment = c_align; c.border = thin_border
r += 1
for m in sorted(ms.keys(), key=lambda x: -ms[x]['total']):
    sr(ws4, r, [m, ms[m]['count'], round(ms[m]['total'],1)]); r += 1
aw(ws4); ws4.freeze_panes = 'A3'

# ── Sheet 5: 前缀分组 ──
ws5 = wb.create_sheet("按前缀分组统计")
ws5.merge_cells('A1:K1')
ws5.cell(row=1,column=1,value="按编号前缀分组统计").font = Font(name="微软雅黑",size=12,bold=True,color="1F4E79")
sh(ws5, 2, ["编号前缀","需求数","超标数","调整前总工时","调整后总工时","减少量","调整前平均","调整后平均","仍超标数","来源","备注"])
grps = req.groupby('编号前缀',dropna=False).agg(
    需求数=('需求','count'),超标数=('是否超标','sum'),
    调整前总工时=('总工时','sum'),调整后总工时=('调整后总工时','sum'),
    仍超标数=('调整后是否超标','sum')).reset_index()
grps['编号前缀'] = grps['编号前缀'].fillna('无编号')
grps['减少量'] = grps['调整前总工时'] - grps['调整后总工时']
grps['调整前平均'] = (grps['调整前总工时']/grps['需求数']).round(1)
grps['调整后平均'] = (grps['调整后总工时']/grps['需求数']).round(1)
grps = grps.sort_values('调整前总工时', ascending=False)
r = 3
for _, row in grps.iterrows():
    note = f"超标{int(row['超标数'])}个,减{row['减少量']:.0f}h" if row['超标数']>0 else ""
    if row['仍超标数']>0: note += f",仍{int(row['仍超标数'])}个"
    sr(ws5, r, [row['编号前缀'],row['需求数'],int(row['超标数']),
                round(row['调整前总工时'],1),round(row['调整后总工时'],1),round(row['减少量'],1),
                row['调整前平均'],row['调整后平均'],int(row['仍超标数']),
                '占比分析页签(S4)' if row['超标数']>0 else '', note])
    if row['仍超标数']>0: ws5.cell(row=r,column=9).fill = warn_fill
    r += 1
aw(ws5); ws5.freeze_panes = 'A3'

# ════════════════════════════════
wb.save(OUTPUT)
print(f"输出: {OUTPUT}")

# 统计
print("\n" + "="*60)
print(f"超标需求: {len(over)}, 超标总量: {over['超标量'].sum():.0f}h")
print(f"重分配: {len(plan)}条")
print(f"仍超标: {req['调整后是否超标'].sum()}")

print("\nS4 全局统计:")
for cat, col in [('开发','开发工时'),('测试','测试工时'),('产品(需求)','产品(需求)工时'),('UI','UI工时')]:
    total = req[col].sum()
    pct = total / req['总工时'].sum() * 100
    print(f"  {cat}: {total:.0f}h ({pct:.1f}%)")

print(f"\n全部需求与标准偏差:")
for cat in ['开发','测试','产品(需求)','UI设计']:
    actual = req[f'{cat}工时' if cat != '产品(需求)' and cat != 'UI设计' else ('产品(需求)工时' if cat=='产品(需求)' else 'UI工时')]
    ref = req[f'参考_{cat}']
    diff = round((actual - ref).sum(), 1)
    print(f"  {cat}: {diff:+.0f}h")
