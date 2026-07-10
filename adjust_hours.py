"""
工时调整分析脚本
- 需求工时 > 450h 的需要重新分配
- 标准配比 开发:测试:产品(需求) = 6:3:2
- 开发任务细分 前端:后端:架构 = 4:5:1 (行业标准配比)
- 生成调整前后对比 Excel
"""
import pandas as pd
import re
import numpy as np
from collections import defaultdict
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# 配置
# ============================================================
THRESHOLD = 450  # 需求工时上限
RATIO_DEV_TEST_PRD = {'开发': 6, '测试': 3, '产品': 2}  # 开发:测试:产品
RATIO_FE_BE_ARCH = {'前端开发': 4, '后端开发': 5, '架构开发': 1}  # 前端:后端:架构
UI_RATIO_IN_PRD = 0.15  # UI占产品工时比例的参考值

INPUT_PATH = '/Users/crystal/软件/workassist/workassist/download\\数据分析072cac2b-5be8-4aad-8c92-dd93564fab40.xlsx'
OUTPUT_PATH = '/Users/crystal/WorkBuddy/禅道任务/工时调整前后对比.xlsx'

# ============================================================
# 数据读取与预处理
# ============================================================
df_raw = pd.read_excel(INPUT_PATH, sheet_name='占比分析_v2')

# 只取明细行（有父任务类型的行）
detail = df_raw[df_raw['父任务类型'].notna() & (df_raw['父任务类型'] != '') & (df_raw['父任务类型'] != '分割行')].copy()

def extract_prefix(name):
    if pd.isna(name) or name == '':
        return None
    m = re.match(r'^(\d+)', str(name).strip())
    if m:
        return m.group(1)
    return None

def extract_name(name):
    """提取需求名称（去掉编号和#id）"""
    if pd.isna(name):
        return ''
    s = str(name).strip()
    s = re.sub(r'\(\#\d+\)$', '', s)
    s = re.sub(r'^\d+\s*[-—–]\s*', '', s)
    return s.strip()

def similarity(a, b):
    """两个字符串的相似度"""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

# 任务类型映射
TASK_TYPE_MAP = {
    '前端开发': '前端开发',
    '前端开发-': '前端开发',
    '后端开发': '后端开发',
    '后端开发-': '后端开发',
    '开发任务': '架构开发',
    '技术设计': '架构开发',
    '测试': '测试',
    '测试-': '测试',
    '需求设计': '产品',
    '需求调研': '产品',
    '需求验收': '产品',
    'UI设计': 'UI设计',
    '运维': '产品',
    '管理事务': '产品',
    '其他': '产品',
    'research': '产品',
}

# 父任务类型标准化
PARENT_TYPE_MAP = {
    '产品': '产品',
    '开发': '开发',
    '测试': '测试',
    'UI': 'UI',
    '未知': '未知',
}

detail['编号前缀'] = detail['相关研发需求'].apply(extract_prefix)
detail['需求简称'] = detail['相关研发需求'].apply(extract_name)
detail['标准化父类型'] = detail['父任务类型'].map(PARENT_TYPE_MAP).fillna('未知')
detail['标准化任务类型'] = detail['任务类型'].map(TASK_TYPE_MAP).fillna('其他')

# ============================================================
# Step 1: 计算需求级汇总（调整前）
# ============================================================
req_summary = detail.groupby('相关研发需求').agg(
    编号前缀=('编号前缀', 'first'),
    需求简称=('需求简称', 'first'),
    总工时=('总计消耗', 'sum'),
    开发工时=('总计消耗', lambda x: x[detail.loc[x.index, '标准化父类型'] == '开发'].sum()),
    测试工时=('总计消耗', lambda x: x[detail.loc[x.index, '标准化父类型'] == '测试'].sum()),
    产品工时=('总计消耗', lambda x: x[detail.loc[x.index, '标准化父类型'] == '产品'].sum()),
    UI工时=('总计消耗', lambda x: x[detail.loc[x.index, '标准化父类型'] == 'UI'].sum()),
).reset_index()

# 开发子类型工时
dev_detail = detail[detail['标准化父类型'] == '开发'].copy()
dev_sub = dev_detail.groupby(['相关研发需求', '标准化任务类型'])['总计消耗'].sum().unstack(fill_value=0)
dev_sub = dev_sub.reindex(columns=['前端开发', '后端开发', '架构开发'], fill_value=0)
req_summary = req_summary.merge(dev_sub.add_prefix('开发_'), left_on='相关研发需求', right_index=True, how='left')

# 筛选 > 450h 的需求
req_summary['是否超标'] = req_summary['总工时'] > THRESHOLD
req_summary['超标量'] = (req_summary['总工时'] - THRESHOLD).clip(lower=0)
over_reqs = req_summary[req_summary['是否超标']].copy()
normal_reqs = req_summary[~req_summary['是否超标']].copy()

print(f"总需求数: {len(req_summary)}")
print(f"超标需求数: {len(over_reqs)}")
print(f"正常需求数: {len(normal_reqs)}")

# ============================================================
# Step 2: 超标工时重分配方案
# ============================================================
# 构建转移方案：超标需求 -> 目标需求
redistribution_plan = []  # [(来源需求, 目标需求, 转移工时, 匹配方式)]

for _, over_req in over_reqs.iterrows():
    src_name = over_req['相关研发需求']
    src_prefix = over_req['编号前缀']
    excess = over_req['超标量']
    remaining_excess = excess
    
    # 优先级1: 同编号前缀，且未超标的需求
    if src_prefix:
        same_prefix_normal = normal_reqs[normal_reqs['编号前缀'] == src_prefix].copy()
        if len(same_prefix_normal) > 0:
            # 计算每个目标还能接收多少（最多到450）
            same_prefix_normal['剩余容量'] = same_prefix_normal['总工时'].apply(lambda x: max(THRESHOLD - x, 0))
            same_prefix_normal = same_prefix_normal[same_prefix_normal['剩余容量'] > 0].sort_values('剩余容量', ascending=False)
            
            for _, target in same_prefix_normal.iterrows():
                if remaining_excess <= 0:
                    break
                move_amount = min(remaining_excess, float(target['剩余容量']))
                redistribution_plan.append((src_name, target['相关研发需求'], move_amount, '同编号前缀'))
                remaining_excess -= move_amount
    
    # 优先级2: 同编号前缀但已超标的需求（分担超标量）
    if remaining_excess > 0 and src_prefix:
        same_prefix_over = over_reqs[(over_reqs['编号前缀'] == src_prefix) & (over_reqs['相关研发需求'] != src_name)].copy()
        if len(same_prefix_over) > 0:
            # 平均分担
            per_req = remaining_excess / len(same_prefix_over)
            for _, target in same_prefix_over.iterrows():
                if remaining_excess <= 0:
                    break
                move_amount = min(remaining_excess, per_req)
                redistribution_plan.append((src_name, target['相关研发需求'], round(move_amount, 1), '同编号分担'))
                remaining_excess -= move_amount
    
    # 优先级3: 名称相似的需求
    if remaining_excess > 0:
        src_simple = extract_name(src_name)
        candidates = normal_reqs.copy()
        candidates['相似度'] = candidates['需求简称'].apply(lambda x: similarity(src_simple, x))
        candidates = candidates[candidates['相似度'] > 0.3].sort_values('相似度', ascending=False)
        
        for _, target in candidates.head(5).iterrows():
            if remaining_excess <= 0:
                break
            target_capacity = max(THRESHOLD - target['总工时'], 0)
            if target_capacity > 0:
                move_amount = min(remaining_excess, target_capacity)
                redistribution_plan.append((src_name, target['相关研发需求'], round(move_amount, 1), f"名称相似({target['相似度']:.0%})"))
                remaining_excess -= move_amount
    
    # 优先级4: 所有需求平均分担剩余
    if remaining_excess > 0:
        all_others = req_summary[req_summary['相关研发需求'] != src_name].copy()
        all_others['剩余容量'] = all_others['总工时'].apply(lambda x: max(THRESHOLD - x, 0))
        all_others = all_others[all_others['剩余容量'] > 0].sort_values('剩余容量', ascending=False)
        per_req = remaining_excess / max(len(all_others), 1)
        for _, target in all_others.head(20).iterrows():
            if remaining_excess <= 0:
                break
            move_amount = min(remaining_excess, per_req, float(target['剩余容量']))
            redistribution_plan.append((src_name, target['相关研发需求'], round(move_amount, 1), '全局分担'))
            remaining_excess -= move_amount

print(f"\n重分配方案条目数: {len(redistribution_plan)}")

# ============================================================
# Step 3: 计算调整后的工时
# ============================================================
# 创建调整后工时字典
adjusted_hours = {req: row['总工时'] for req, row in req_summary.set_index('相关研发需求').iterrows()}

# 应用转移
for src, tgt, amount, method in redistribution_plan:
    adjusted_hours[src] = adjusted_hours.get(src, 0) - amount
    adjusted_hours[tgt] = adjusted_hours.get(tgt, 0) + amount

# 确保不超过450（来源需求）
for src, tgt, amount, method in redistribution_plan:
    if adjusted_hours[src] > THRESHOLD:
        diff = adjusted_hours[src] - THRESHOLD
        adjusted_hours[src] = THRESHOLD
    adjusted_hours[src] = max(adjusted_hours[src], 0)

req_summary['调整后总工时'] = req_summary['相关研发需求'].map(adjusted_hours)
req_summary['工时变化'] = req_summary['调整后总工时'] - req_summary['总工时']
req_summary['调整后是否超标'] = req_summary['调整后总工时'] > THRESHOLD

# ============================================================
# Step 4: 标准配比参考值计算
# ============================================================
ratio_sum = sum(RATIO_DEV_TEST_PRD.values())
ratio_fe_be_arch_sum = sum(RATIO_FE_BE_ARCH.values())

def calc_standard_ratios(total_hours):
    """根据标准配比计算各类型工时"""
    dev_hours = total_hours * RATIO_DEV_TEST_PRD['开发'] / ratio_sum
    test_hours = total_hours * RATIO_DEV_TEST_PRD['测试'] / ratio_sum
    prd_hours = total_hours * RATIO_DEV_TEST_PRD['产品'] / ratio_sum
    
    fe_hours = dev_hours * RATIO_FE_BE_ARCH['前端开发'] / ratio_fe_be_arch_sum
    be_hours = dev_hours * RATIO_FE_BE_ARCH['后端开发'] / ratio_fe_be_arch_sum
    arch_hours = dev_hours * RATIO_FE_BE_ARCH['架构开发'] / ratio_fe_be_arch_sum
    
    ui_hours = prd_hours * UI_RATIO_IN_PRD
    pure_prd_hours = prd_hours - ui_hours
    
    return {
        '开发': round(dev_hours, 1),
        '测试': round(test_hours, 1),
        '产品': round(prd_hours, 1),
        'UI': round(ui_hours, 1),
        '前端开发': round(fe_hours, 1),
        '后端开发': round(be_hours, 1),
        '架构开发': round(arch_hours, 1),
    }

# 为每个超标需求计算标准配比参考值
for idx, row in req_summary.iterrows():
    target_total = min(row['调整后总工时'], THRESHOLD)
    std = calc_standard_ratios(target_total)
    for k, v in std.items():
        req_summary.at[idx, f'参考_{k}'] = v

print(f"\n调整后仍超标的需求: {req_summary['调整后是否超标'].sum()}")

# ============================================================
# Step 5: 生成对比 Excel
# ============================================================
wb = Workbook()

# --- 颜色和样式 ---
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
sub_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sub_header_font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")

over_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 超标红色
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 正常绿色
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 警告黄色
adjusted_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
normal_font = Font(name="微软雅黑", size=9)
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols, fill=None, font=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill or header_fill
        cell.font = font or header_font
        cell.alignment = center_align
        cell.border = thin_border

def style_data_row(ws, row, cols, is_over=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = center_align if c > 1 else left_align
        cell.border = thin_border

def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_w
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)) * 1.2 + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_w)

# ============================================================
# Sheet 1: 需求级对比总览
# ============================================================
ws1 = wb.active
ws1.title = "需求级调整对比"

# 标题行
ws1.merge_cells('A1:R1')
ws1.cell(row=1, column=1, value="工时调整前后对比总览 — 标准配比 开发:测试:产品=6:3:2 | 前端:后端:架构=4:5:1").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

headers1 = [
    "序号", "相关研发需求", "编号前缀",
    "调整前总工时", "超标量",
    "调整后总工时", "工时变化",
    # 调整前分布
    "调整前-开发", "调整前-测试", "调整前-产品", "调整前-UI",
    # 调整前开发细分
    "调整前-前端", "调整前-后端", "调整前-架构",
    # 标准配比参考值
    "参考-开发(54.5%)", "参考-测试(27.3%)", "参考-产品(18.2%)",
    "参考-前端", "参考-后端", "参考-架构",
    # 差异分析
    "开发偏差", "测试偏差", "产品偏差",
    "状态"
]

for i, h in enumerate(headers1, 1):
    ws1.cell(row=2, column=i, value=h)
style_header(ws1, 2, len(headers1))

# 数据行 - 先超标需求，再正常需求
row_num = 3
over_sorted = req_summary[req_summary['是否超标']].sort_values('总工时', ascending=False)

for _, r in over_sorted.iterrows():
    ws1.cell(row=row_num, column=1, value=row_num - 2)
    ws1.cell(row=row_num, column=2, value=r['相关研发需求'])
    ws1.cell(row=row_num, column=3, value=r['编号前缀'] or '无')
    ws1.cell(row=row_num, column=4, value=r['总工时'])
    ws1.cell(row=row_num, column=5, value=r['超标量'])
    ws1.cell(row=row_num, column=6, value=r['调整后总工时'])
    ws1.cell(row=row_num, column=7, value=r['工时变化'])
    ws1.cell(row=row_num, column=8, value=r['开发工时'])
    ws1.cell(row=row_num, column=9, value=r['测试工时'])
    ws1.cell(row=row_num, column=10, value=r['产品工时'])
    ws1.cell(row=row_num, column=11, value=r['UI工时'])
    ws1.cell(row=row_num, column=12, value=r['开发_前端开发'])
    ws1.cell(row=row_num, column=13, value=r['开发_后端开发'])
    ws1.cell(row=row_num, column=14, value=r['开发_架构开发'])
    ws1.cell(row=row_num, column=15, value=r['参考_开发'])
    ws1.cell(row=row_num, column=16, value=r['参考_测试'])
    ws1.cell(row=row_num, column=17, value=r['参考_产品'])
    ws1.cell(row=row_num, column=18, value=r['参考_前端开发'])
    ws1.cell(row=row_num, column=19, value=r['参考_后端开发'])
    ws1.cell(row=row_num, column=20, value=r['参考_架构开发'])
    
    dev_diff = r['开发工时'] - r['参考_开发']
    test_diff = r['测试工时'] - r['参考_测试']
    prd_diff = r['产品工时'] - r['参考_产品']
    ws1.cell(row=row_num, column=21, value=round(dev_diff, 1))
    ws1.cell(row=row_num, column=22, value=round(test_diff, 1))
    ws1.cell(row=row_num, column=23, value=round(prd_diff, 1))
    
    if r['调整后是否超标']:
        ws1.cell(row=row_num, column=24, value="仍需调整")
    else:
        ws1.cell(row=row_num, column=24, value="已达标")
    
    style_data_row(ws1, row_num, len(headers1), is_over=True)
    # 超标行标红
    if r['是否超标']:
        for c in [4, 5]:
            ws1.cell(row=row_num, column=c).fill = over_fill
    if not r['调整后是否超标']:
        ws1.cell(row=row_num, column=6).fill = ok_fill
    
    row_num += 1

auto_width(ws1)
ws1.freeze_panes = 'C3'

# ============================================================
# Sheet 2: 重分配明细
# ============================================================
ws2 = wb.create_sheet("工时重分配明细")

ws2.merge_cells('A1:F1')
ws2.cell(row=1, column=1, value="超标工时重分配明细").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

headers2 = ["序号", "来源需求", "来源前缀", "转移工时(h)", "目标需求", "目标前缀", "匹配方式", "来源调整后", "目标调整后"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=2, column=i, value=h)
style_header(ws2, 2, len(headers2))

row_num = 3
for idx, (src, tgt, amount, method) in enumerate(redistribution_plan, 1):
    ws2.cell(row=row_num, column=1, value=idx)
    ws2.cell(row=row_num, column=2, value=src)
    ws2.cell(row=row_num, column=3, value=extract_prefix(src) or '无')
    ws2.cell(row=row_num, column=4, value=amount)
    ws2.cell(row=row_num, column=5, value=tgt)
    ws2.cell(row=row_num, column=6, value=extract_prefix(tgt) or '无')
    ws2.cell(row=row_num, column=7, value=method)
    ws2.cell(row=row_num, column=8, value=round(adjusted_hours.get(src, 0), 1))
    ws2.cell(row=row_num, column=9, value=round(adjusted_hours.get(tgt, 0), 1))
    style_data_row(ws2, row_num, len(headers2))
    row_num += 1

# 按匹配方式汇总
ws2.cell(row=row_num + 1, column=1, value="汇总").font = Font(name="微软雅黑", size=10, bold=True)
method_summary = defaultdict(lambda: {'count': 0, 'total': 0})
for src, tgt, amount, method in redistribution_plan:
    method_summary[method]['count'] += 1
    method_summary[method]['total'] += amount

for i, (method, stats) in enumerate(sorted(method_summary.items())):
    ws2.cell(row=row_num + 2 + i, column=1, value=method)
    ws2.cell(row=row_num + 2 + i, column=2, value=f"{stats['count']}条")
    ws2.cell(row=row_num + 2 + i, column=3, value=f"{stats['total']:.1f}h")

auto_width(ws2)
ws2.freeze_panes = 'A3'

# ============================================================
# Sheet 3: 标准配比参考
# ============================================================
ws3 = wb.create_sheet("标准配比参考")

ws3.merge_cells('A1:H1')
ws3.cell(row=1, column=1, value="标准工时配比参考表 — 开发:测试:产品 = 6:3:2 | 前端:后端:架构 = 4:5:1").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

# 配比说明
ws3.merge_cells('A3:H3')
ws3.cell(row=3, column=1, value="配比说明").font = Font(name="微软雅黑", size=11, bold=True)
ws3.cell(row=3, column=1).fill = sub_header_fill

ratio_info = [
    ["层级", "类型", "配比", "占比", "说明"],
    ["一级", "开发", "6", "54.55%", "所有开发相关任务（前端+后端+架构+技术设计）"],
    ["一级", "测试", "3", "27.27%", "功能测试、集成测试、回归测试等"],
    ["一级", "产品(需求)", "2", "18.18%", "需求设计、调研、验收、UI设计、运维、管理事务等"],
    ["", "", "", "", ""],
    ["二级(开发内)", "前端开发", "4", "40%", "前端页面、组件、交互开发"],
    ["二级(开发内)", "后端开发", "5", "50%", "后端接口、逻辑、数据处理"],
    ["二级(开发内)", "架构开发", "1", "10%", "架构设计、技术方案、开发任务"],
    ["", "", "", "", ""],
    ["三级(产品内)", "纯产品", "85%", "—", "需求设计、调研、验收、管理事务"],
    ["三级(产品内)", "UI设计", "15%", "—", "界面设计、交互设计"],
]

for i, row_data in enumerate(ratio_info, 4):
    for j, val in enumerate(row_data):
        ws3.cell(row=i, column=j+1, value=val)
        ws3.cell(row=i, column=j+1).font = normal_font
        ws3.cell(row=i, column=j+1).border = thin_border

# 不同总工时下的参考值表
ws3.cell(row=16, column=1, value="不同需求总工时下的标准工时参考值").font = Font(name="微软雅黑", size=11, bold=True)
ws3.cell(row=16, column=1).fill = sub_header_fill
ws3.merge_cells('A16:H16')

ref_headers = ["需求总工时", "开发(54.5%)", "测试(27.3%)", "产品(18.2%)", "前端开发(21.8%)", "后端开发(27.3%)", "架构开发(5.5%)", "UI设计(2.7%)"]
for i, h in enumerate(ref_headers, 1):
    ws3.cell(row=17, column=i, value=h)
style_header(ws3, 17, len(ref_headers))

sample_hours = [100, 200, 300, 400, 450, 500, 600, 800, 1000]
for idx, total in enumerate(sample_hours, 18):
    std = calc_standard_ratios(min(total, THRESHOLD))
    ws3.cell(row=idx, column=1, value=total)
    ws3.cell(row=idx, column=2, value=std['开发'])
    ws3.cell(row=idx, column=3, value=std['测试'])
    ws3.cell(row=idx, column=4, value=std['产品'])
    ws3.cell(row=idx, column=5, value=std['前端开发'])
    ws3.cell(row=idx, column=6, value=std['后端开发'])
    ws3.cell(row=idx, column=7, value=std['架构开发'])
    ws3.cell(row=idx, column=8, value=std['UI'])
    style_data_row(ws3, idx, len(ref_headers))

auto_width(ws3)

# ============================================================
# Sheet 4: 按编号前缀汇总
# ============================================================
ws4 = wb.create_sheet("按编号前缀汇总")

ws4.merge_cells('A1:I1')
ws4.cell(row=1, column=1, value="按编号前缀分组 — 工时分布与超标情况").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

prefix_headers = ["编号前缀", "需求数量", "超标需求数", "调整前总工时", "超标总量", "调整后总工时", "平均工时", "仍超标数", "备注"]
for i, h in enumerate(prefix_headers, 1):
    ws4.cell(row=2, column=i, value=h)
style_header(ws4, 2, len(prefix_headers))

prefix_groups = req_summary.groupby('编号前缀', dropna=False).agg(
    需求数量=('相关研发需求', 'count'),
    超标需求数=('是否超标', 'sum'),
    调整前总工时=('总工时', 'sum'),
    超标总量=('超标量', 'sum'),
    调整后总工时=('调整后总工时', 'sum'),
    仍超标数=('调整后是否超标', 'sum'),
).reset_index()

prefix_groups['编号前缀'] = prefix_groups['编号前缀'].fillna('无编号')
prefix_groups['平均工时'] = (prefix_groups['调整前总工时'] / prefix_groups['需求数量']).round(1)
prefix_groups = prefix_groups.sort_values('调整前总工时', ascending=False)

for idx, (_, row) in enumerate(prefix_groups.iterrows(), 3):
    ws4.cell(row=idx, column=1, value=row['编号前缀'])
    ws4.cell(row=idx, column=2, value=row['需求数量'])
    ws4.cell(row=idx, column=3, value=row['超标需求数'])
    ws4.cell(row=idx, column=4, value=round(row['调整前总工时'], 1))
    ws4.cell(row=idx, column=5, value=round(row['超标总量'], 1))
    ws4.cell(row=idx, column=6, value=round(row['调整后总工时'], 1))
    ws4.cell(row=idx, column=7, value=row['平均工时'])
    ws4.cell(row=idx, column=8, value=row['仍超标数'])
    if row['超标需求数'] > 0:
        ws4.cell(row=idx, column=9, value=f"需重分配{row['超标总量']:.0f}h")
    style_data_row(ws4, idx, len(prefix_headers))

auto_width(ws4)
ws4.freeze_panes = 'A3'

# ============================================================
# Sheet 5: 开发任务细分对比
# ============================================================
ws5 = wb.create_sheet("开发任务细分对比")

ws5.merge_cells('A1:L1')
ws5.cell(row=1, column=1, value="开发任务细分 — 前端:后端:架构 实际 vs 参考(4:5:1)").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

dev_headers = [
    "序号", "相关研发需求",
    "开发总工时", "前端实际", "后端实际", "架构实际",
    "前端参考(40%)", "后端参考(50%)", "架构参考(10%)",
    "前端偏差", "后端偏差", "架构偏差"
]
for i, h in enumerate(dev_headers, 1):
    ws5.cell(row=2, column=i, value=h)
style_header(ws5, 2, len(dev_headers))

row_num = 3
for _, r in over_sorted.iterrows():
    dev_total = r['开发工时']
    if dev_total > 0:
        ref_fe = round(dev_total * 0.4, 1)
        ref_be = round(dev_total * 0.5, 1)
        ref_arch = round(dev_total * 0.1, 1)
        ws5.cell(row=row_num, column=1, value=row_num - 2)
        ws5.cell(row=row_num, column=2, value=r['相关研发需求'])
        ws5.cell(row=row_num, column=3, value=dev_total)
        ws5.cell(row=row_num, column=4, value=r['开发_前端开发'])
        ws5.cell(row=row_num, column=5, value=r['开发_后端开发'])
        ws5.cell(row=row_num, column=6, value=r['开发_架构开发'])
        ws5.cell(row=row_num, column=7, value=ref_fe)
        ws5.cell(row=row_num, column=8, value=ref_be)
        ws5.cell(row=row_num, column=9, value=ref_arch)
        ws5.cell(row=row_num, column=10, value=round(r['开发_前端开发'] - ref_fe, 1))
        ws5.cell(row=row_num, column=11, value=round(r['开发_后端开发'] - ref_be, 1))
        ws5.cell(row=row_num, column=12, value=round(r['开发_架构开发'] - ref_arch, 1))
        style_data_row(ws5, row_num, len(dev_headers))
        row_num += 1

auto_width(ws5)
ws5.freeze_panes = 'C3'

# ============================================================
# 保存
# ============================================================
wb.save(OUTPUT_PATH)
print(f"\n文件已保存: {OUTPUT_PATH}")
print(f"Sheets: {wb.sheetnames}")

# ============================================================
# 关键统计输出
# ============================================================
print("\n" + "="*60)
print("关键统计")
print("="*60)
print(f"超标需求数: {len(over_reqs)}")
print(f"超标总工时: {over_reqs['超标量'].sum():.1f}h")
print(f"重分配方案条目: {len(redistribution_plan)}")
print(f"重分配总工时: {sum(a for _,_,a,_ in redistribution_plan):.1f}h")
print(f"调整后仍超标需求数: {req_summary['调整后是否超标'].sum()}")

# 按匹配方式统计
print("\n匹配方式统计:")
for method, stats in sorted(method_summary.items()):
    print(f"  {method}: {stats['count']}条, {stats['total']:.1f}h")

# 标准差分析
print("\n偏离标准配比最大的需求 (TOP 5):")
req_summary['开发偏差'] = abs(req_summary['开发工时'] - req_summary['参考_开发'])
top5 = req_summary[req_summary['是否超标']].nlargest(5, '开发偏差')
for _, r in top5.iterrows():
    pct = r['开发工时'] / r['总工时'] * 100 if r['总工时'] > 0 else 0
    target_pct = 600/11  # 54.55%
    print(f"  {r['相关研发需求'][:40]}... 开发占{pct:.1f}%(目标{target_pct:.1f}%), 偏差{r['开发偏差']:.1f}h")
