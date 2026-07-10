#!/usr/bin/env python3
"""需求工时审计：7月1日至今，按需求维度4项检查"""
import openpyxl, json, re, os
from datetime import date, datetime
from collections import defaultdict

# ===== 配置 =====
UI_FILE = '/Users/crystal/Downloads/任务完成汇总表UI (1).xlsx'
DEV_FILE = '/Users/crystal/Downloads/任务完成汇总表开发 (1).xlsx'
HOURS_FILE = '/Users/crystal/WorkBuddy/禅道任务/authoritative_hours.json'
OUTPUT = '/Users/crystal/WorkBuddy/禅道任务/requirement_audit.json'

START_DATE = date(2026, 7, 1)
END_DATE = date(2026, 7, 8)

# 任务类型识别规则（按优先级）
TASK_TYPE_PATTERNS = [
    (r'【UI设计】|【UI\d*】|【UI】', 'UI设计'),
    (r'【UX设计】|【UX\d*】|【UX】', 'UX设计'),
    (r'【前端】', '前端开发'),
    (r'【后端】', '后端开发'),
    (r'【产品\d*】|【产品】', '产品'),
    (r'【测试】', '测试'),
    (r'需求调研|需求设计|需求验收|需求分析|需求沟通|需求评审', '产品'),
]

# 部门→默认任务类型
DEPT_DEFAULT_TYPE = {
    '视觉设计室': 'UI设计',
    '开发七室': '后端开发',
    '软件测试室': '测试',
}

# 部门→人员大角色（用于检查4）
DEPT_EXPECTED_ROLE = {
    '视觉设计室': '设计',
    '开发七室': '开发',
    '软件测试室': '测试',
}

EXCLUDE_PERSONS = {'苏方进', '黎思斯', '张旭', '余晨伟', '蔡能', '程迎娣', '胡丹'}

TYPE_TO_CATEGORY = {
    'UI设计': '设计', 'UX设计': '设计',
    '前端开发': '开发', '后端开发': '开发',
    '产品': '产品', '测试': '测试',
}

AUDIT_MAX_TOTAL = 450
AUDIT_MIN_TOTAL = 40
AUDIT_BACKEND_MAX = 245

# ===== 工具函数 =====
def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    try: return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except: return None

def classify_task(task_name, person_dept):
    """任务类型识别：名称前缀优先，否则按人员部门推断"""
    for pattern, ttype in TASK_TYPE_PATTERNS:
        if re.search(pattern, task_name):
            return ttype
    return DEPT_DEFAULT_TYPE.get(person_dept, '未分类')

def extract_requirement(task_name, execution):
    """从任务名称提取需求编号和名称，无编号的用执行兜底"""
    # 1. 去掉【】【】标签
    name = re.sub(r'^【[^】]+】\s*', '', task_name)
    name = re.sub(r'^\[[^\]]+\]\s*', '', name)
    name = name.strip()
    
    # 2. 匹配4位需求编号 + 描述（如 "1222-收银台v1.7.0-收银台支持..."）
    m = re.match(r'^(\d{3,4})[-\s]+(.+?)$', name)
    if m:
        num = m.group(1)
        desc = m.group(2).strip()
        # 取第一个有意义的词段
        first_segment = desc.split('-')[0].strip()
        # 提取前4个中文字作为标识（够区分即可）
        cn = re.findall(r'[\u4e00-\u9fa5]+', first_segment)
        cn_str = ''.join(cn)
        short = cn_str[:6] if cn_str else first_segment[:6]
        return (num, short)
    
    # 3. 匹配纯4位数字（无描述）
    m = re.match(r'^(\d{3,4})\b', name)
    if m:
        return (m.group(1), m.group(1))
    
    # 4. 无编号的任务按执行兜底
    exec_short = execution.replace('/【2026】', '') if execution else '未归类'
    return (exec_short, '')  # 名称留空，页面显示时只用编号列

# ===== 加载数据 =====
def load_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Sheet1']
    rows = []
    cp, cproj, cexec = None, None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        if i == 0: continue
        person = row[0] if row[0] else cp; cp = person
        project = row[1] if row[1] else cproj; cproj = project
        execution = row[2] if row[2] else cexec; cexec = execution
        rows.append({
            '完成者': person, '所属项目': project, '所属执行': execution,
            '编号': row[3], '任务名称': row[4], 'P': row[5],
            '预计开始': row[6], '实际开始': row[7], '截止日期': row[8],
            '实际完成': row[9], '延期(天)': row[10],
            '最初预计': float(row[11]) if row[11] else 0,
            '任务消耗工时': float(row[12]) if row[12] else 0,
            '总任务数': row[13], '执行总消耗': row[14], '用户总消耗': row[15],
        })
    return rows

# 加载人员部门映射
with open(HOURS_FILE) as f:
    hours_data = json.load(f)
person_dept_map = {}
for name, info in hours_data.items():
    if info.get('dept'):
        person_dept_map[name] = info['dept']

print("加载数据...")
ui_rows = load_excel(UI_FILE)
dev_rows = load_excel(DEV_FILE)
all_rows = ui_rows + dev_rows

# 过滤7月1日-今日
filtered = [r for r in all_rows if (d := parse_date(r['实际完成'])) and START_DATE <= d <= END_DATE]
print(f"7/1-7/8 任务: {len(filtered)} 条, 涉及 {len(set(r['完成者'] for r in filtered))} 人")

# ===== 按需求分组审计 =====
# 先用执行作为兜底分组，优先按任务名中的需求标识
req_map = defaultdict(lambda: {
    '需求名称': '', '项目': '', '所属执行': '', '任务列表': [],
    '类型工时': defaultdict(float), '大类工时': defaultdict(float),
})

for r in filtered:
    if r['完成者'] in EXCLUDE_PERSONS:
        continue
    person = r['完成者']
    dept = person_dept_map.get(person, '未知')
    task_type = classify_task(r['任务名称'], dept)
    category = TYPE_TO_CATEGORY.get(task_type, '开发')
    
    # 提取需求标识：(编号, 名称)
    req_id, req_name = extract_requirement(r['任务名称'], r['所属执行'])
    req_key = f"{req_id}_{req_name}"
    
    e = req_map[req_key]
    e['需求编号'] = req_id
    e['需求名称'] = req_name
    e['项目'] = r['所属项目'] or e['项目']
    # 记录所有涉及执行
    ex = r['所属执行']
    if ex and ex not in e.setdefault('执行列表', []):
        e['执行列表'].append(ex)
    e['任务列表'].append({
        '完成者': person, '部门': dept,
        '任务名称': r['任务名称'], '任务类型': task_type,
        '大类': category, '工时': r['任务消耗工时'], '编号': r['编号'],
        '所属执行': r['所属执行'],
    })
    e['类型工时'][task_type] += r['任务消耗工时']
    e['大类工时'][category] += r['任务消耗工时']

print(f"  按需求分组: {len(req_map)} 个需求")

# ===== 审计 =====
audit_results = []
for req_key, e in sorted(req_map.items()):
    total_hours = sum(t['工时'] for t in e['任务列表'])
    cat_hours = dict(e['大类工时'])
    
    # 后端开发最高单任务
    backend_tasks = [t for t in e['任务列表'] if t['任务类型'] == '后端开发']
    backend_max = max((t['工时'] for t in backend_tasks), default=0)
    # 也检查未分类但按部门归到开发的
    if backend_max == 0:
        dev_tasks = [t for t in e['任务列表'] if t['大类'] == '开发']
        backend_max = max((t['工时'] for t in dev_tasks), default=0)
    
    warnings = []
    if total_hours > AUDIT_MAX_TOTAL:
        warnings.append(f'总工时{total_hours:.1f}h 超过{AUDIT_MAX_TOTAL}h上限')
    if total_hours < AUDIT_MIN_TOTAL:
        warnings.append(f'总工时{total_hours:.1f}h 低于{AUDIT_MIN_TOTAL}h下限')
    if backend_max > AUDIT_BACKEND_MAX:
        warnings.append(f'开发单任务最高{backend_max:.1f}h 超过{AUDIT_BACKEND_MAX}h上限')
    for cat in ['产品', '开发', '测试']:
        if cat_hours.get(cat, 0) == 0:
            warnings.append(f'{cat}工时为0')
        elif cat == '测试' and cat_hours.get(cat, 0) < 4:
            warnings.append(f'测试工时仅{cat_hours[cat]:.1f}h（偏低）')
    
    audit_results.append({
        '需求编号': e['需求编号'],
        '需求名称': e['需求名称'],
        '所属执行': ', '.join(e.get('执行列表', [])),
        '项目': e['项目'],
        '总工时': round(total_hours, 1), '任务数': len(e['任务列表']),
        '类型工时': {k: round(v, 1) for k, v in sorted(e['类型工时'].items())},
        '大类工时': {k: round(v, 1) for k, v in sorted(cat_hours.items())},
        '开发单任务最高': round(backend_max, 1),
        '预警': warnings, '状态': 'warn' if warnings else 'ok',
    })

# ===== 人员类型审计 =====
person_audit = []
for person in sorted(set(r['完成者'] for r in filtered if r['完成者'] not in EXCLUDE_PERSONS)):
    dept = person_dept_map.get(person, '未知')
    expected_role = DEPT_EXPECTED_ROLE.get(dept, '未知')
    
    person_tasks = [r for r in filtered if r['完成者'] == person]
    type_hours = defaultdict(float)
    category_hours = defaultdict(float)
    for r in person_tasks:
        tt = classify_task(r['任务名称'], dept)
        cat = TYPE_TO_CATEGORY.get(tt, '开发')
        type_hours[tt] += r['任务消耗工时']
        category_hours[cat] += r['任务消耗工时']
    
    # 检查任务类型是否与部门匹配
    total = sum(type_hours.values())
    mismatches = []
    for cat, hrs in category_hours.items():
        if cat != expected_role and hrs > 0:
            pct = round(hrs / total * 100, 1) if total > 0 else 0
            if pct > 30:  # 超过30%工时不在预期范畴才告警
                mismatches.append(f'{cat}类占{pct}%（预期{expected_role}）')
    
    person_audit.append({
        '人员': person, '部门': dept,
        '工时': round(total, 1),
        '任务类型分布': {k: round(v, 1) for k, v in sorted(type_hours.items())},
        '异常': mismatches, '状态': 'warn' if mismatches else 'ok',
    })

# ===== 输出 =====
warn_count = sum(1 for a in audit_results if a['状态'] == 'warn')
p_warn_count = sum(1 for p in person_audit if p['状态'] == 'warn')

output_data = {
    'summary': {
        '统计时间': f'{START_DATE} ~ {END_DATE}',
        '数据来源': '任务完成汇总表(UI+开发)',
        '需求总数': len(audit_results), '预警需求数': warn_count,
        '涉及人员数': len(person_audit), '人员类型异常数': p_warn_count,
    },
    'rules': {
        '总工时上限': f'{AUDIT_MAX_TOTAL}h', '总工时下限': f'{AUDIT_MIN_TOTAL}h',
        '开发单任务上限': f'{AUDIT_BACKEND_MAX}h',
        '必需角色': '产品、开发、测试',
    },
    '需求审计': audit_results,
    '人员审计': person_audit,
}

with open(OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"\n需求审计: {len(audit_results)}个需求, 预警{warn_count}个")
for a in audit_results:
    print(f"  {'⚠️' if a['状态']=='warn' else '✅'} {a['需求编号']}-{a['需求名称']} ({a['所属执行']}): {a['总工时']}h, {a['大类工时']}, 预警={a['预警']}")

print(f"\n人员审计: {len(person_audit)}人, 类型异常{p_warn_count}人")
for p in person_audit:
    flag = '⚠️' if p['状态'] == 'warn' else '✅'
    print(f"  {flag} {p['人员']}({p['部门']}): {p['任务类型分布']} {p['异常']}")

print(f"\n写入: {OUTPUT}")
