#!/usr/bin/env python3
"""将禅道不符合项整改报告CSV转为xlsx，并解析用户ID为真实姓名"""
import json, csv, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = '/Users/crystal/WorkBuddy/禅道任务/禅道问题检查'
DATA_DIR = os.path.join(BASE, 'data_cache')

# Load name map
with open(os.path.join(DATA_DIR, 'user_name_map.json')) as f:
    name_map = json.load(f)

def resolve_name(val):
    if not val: return ''
    v = str(val).strip()
    if v in name_map:
        return f'{name_map[v]}({v})'
    if v.lower() in name_map:
        return f'{name_map[v.lower()]}({v})'
    m = re.match(r'(.+?)\((.+?)\)', v)
    if m:
        return v
    return v

def read_xlsx_rows(filepath):
    """Read rows from xlsx using zipfile+xml"""
    import zipfile, xml.etree.ElementTree as ET
    zf = zipfile.ZipFile(filepath)
    ss = zf.read('xl/sharedStrings.xml')
    root = ET.fromstring(ss)
    ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    strings = [t.text or '' for t in root.findall('.//s:t', ns)]
    
    sheet = zf.read('xl/worksheets/sheet1.xml')
    root = ET.fromstring(sheet)
    rows = []
    for row_elem in root.findall('.//s:row', ns):
        cells = ['' for _ in range(20)]
        for c in row_elem.findall('s:c', ns):
            v = c.find('s:v', ns)
            val = (v.text or '') if v is not None else ''
            t = c.get('t')
            if t == 's' and val:
                try:
                    idx = int(val)
                    if idx < len(strings):
                        val = strings[idx]
                except: pass
            # Get column letter/number
            ref = c.get('r', '')
            col_letter = ''.join(ch for ch in ref if ch.isalpha())
            col_idx = 0
            for ch in col_letter:
                col_idx = col_idx * 26 + (ord(ch.upper()) - ord('A') + 1)
            col_idx -= 1
            if col_idx < len(cells):
                cells[col_idx] = val
            else:
                cells.extend([''] * (col_idx - len(cells) + 1))
                cells[col_idx] = val
        rows.append(cells)
    zf.close()
    return rows

# Read data
xlsx_path = os.path.join(BASE, '禅道不符合项整改报告_2026-06-08.xlsx')
csv_path = os.path.join(BASE, '禅道不符合项整改报告_2026-06-08.csv')

if os.path.exists(csv_path):
    with open(csv_path, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)
else:
    rows = read_xlsx_rows(xlsx_path)

header = rows[0]
data_rows = rows[1:]

print(f'Header: {len(header)} cols, {len(data_rows)} data rows')

# Find indices
idx_creator = header.index('由谁创建')
idx_finisher = header.index('由谁完成')

# Resolve names
resolved_count = 0
unresolved = set()
for row in data_rows:
    while len(row) <= max(idx_creator, idx_finisher):
        row.append('')
    old_c = row[idx_creator]
    old_f = row[idx_finisher]
    row[idx_creator] = resolve_name(old_c)
    row[idx_finisher] = resolve_name(old_f)
    if old_c and old_c != row[idx_creator]:
        resolved_count += 1
    if old_c and old_c not in name_map and old_c.lower() not in name_map:
        unresolved.add(old_c)
    if old_f and old_f not in name_map and old_f.lower() not in name_map:
        unresolved.add(old_f)

print(f'Resolved: {resolved_count} creators')
print(f'Unresolved users: {len(unresolved)} -> {sorted(unresolved)[:20]}')

# Write xlsx
wb = Workbook()
ws = wb.active
ws.title = '不符合项整改报告'

header_font = Font(bold=True, size=11)
header_fill = PatternFill('solid', fgColor='D9E1F2')
p0_fill = PatternFill('solid', fgColor='FFC7CE')
p1_fill = PatternFill('solid', fgColor='FFEB9C')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for ci, h in enumerate(header):
    cell = ws.cell(row=1, column=ci+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for ri, row in enumerate(data_rows):
    pri = row[0] if row else ''
    for ci, val in enumerate(row):
        cell = ws.cell(row=ri+2, column=ci+1, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if ci == 0:
            if 'P0' in str(val):
                cell.fill = p0_fill
            elif 'P1' in str(val):
                cell.fill = p1_fill

col_widths = [8, 12, 10, 22, 24, 50, 50, 38, 12, 14, 12, 12, 12, 12, 8, 20, 20, 20, 20, 58]
for ci in range(len(header)):
    letter = ''
    n = ci
    while n >= 0:
        letter = chr(65 + n % 26) + letter
        n = n // 26 - 1
    if ci < len(col_widths):
        ws.column_dimensions[letter].width = col_widths[ci]

ws.freeze_panes = 'A2'
ws.auto_filter.ref = 'A1:S1'

wb.save(xlsx_path)
print(f'Done: {xlsx_path}')
