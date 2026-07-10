#!/usr/bin/env python3
"""从禅道API获取7/1-7/8任务，校验每个人的任务类型（按任务类型.xlsx，产品岗按名称匹配）"""
import json, urllib.request, ssl, re
import openpyxl
from collections import defaultdict

ssl_ctx = ssl.create_default_context()
ssl_ctx.check_hostname = False
ssl_ctx.verify_mode = ssl.CERT_NONE

TOKEN = "38b70ec837ebe462f78028095b861057"
BASE = "https://ztpm.gree.com:8888/api.php/v2"

def api_get(path):
    url = f"{BASE}{path}"
    req = urllib.request.Request(url, headers={"token": TOKEN})
    try:
        with urllib.request.urlopen(req, context=ssl_ctx, timeout=30) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        print(f"  ERROR: {path} -> {e}")
        return None

# ===== 1. 读取任务类型.xlsx =====
print("Step 1: 建立禅道账号→姓名映射...")
wb = openpyxl.load_workbook('/Users/crystal/WorkBuddy/禅道任务/任务类型.xlsx', data_only=True)
ws = wb['任务类型']

person_expected = {}  # name -> {role, type_code or type_options}
account_to_name = {}

# 产品岗4人：多类型，按任务名称匹配
MULTI_TYPE_PERSONS = {'徐中然', '何曦宇', '连朔', '蔡泽平'}

# 任务名称→类型匹配规则（产品岗）
PRODUCT_TASK_PATTERNS = [
    (r'需求验收|验收', 'ab3_request_check'),     # 先匹配更具体的
    (r'需求调研|调研', 'ab1_needs_research'),
    (r'需求设计|需求评审|设计', 'ab2_request_des'),
]

def parse_type_codes(type_str):
    """解析类型字符串，返回 [(名称, 代码), ...]"""
    codes = re.findall(r'([^(]+)\(#([^)]+)\)', type_str)
    return codes  # [(name, code), ...]

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    name = str(row[0]).strip() if row[0] else None
    role = str(row[1]).strip() if row[1] else None
    type_str = str(row[2]).strip() if row[2] else None
    account_str = str(row[3]).strip() if row[3] else None
    
    if not name or name == 'None':
        continue
    
    type_codes = parse_type_codes(type_str)
    
    if name in MULTI_TYPE_PERSONS and len(type_codes) > 1:
        # 多类型：存储允许的类型列表，按任务名匹配
        person_expected[name] = {
            'role': role,
            'multi_type': True,
            'allowed_types': set(c[1] for c in type_codes),
            'type_options': type_codes,
        }
    else:
        # 单类型
        type_code = type_codes[0][1] if type_codes else type_str
        person_expected[name] = {
            'role': role, 'type_code': type_code, 'multi_type': False,
            'type_options': type_codes,
        }
    
    # 提取禅道账号
    if account_str:
        m = re.search(r'#([^)]+)', account_str)
        if m:
            zt_account = m.group(1)
            account_to_name[zt_account] = name
            account_to_name[zt_account.lower()] = name

print(f"  共 {len(person_expected)} 人, {len(account_to_name)} 个账号映射")
for name, info in person_expected.items():
    if info.get('multi_type'):
        print(f"    {name}({info['role']}): 多类型 -> {info['allowed_types']}")
    else:
        print(f"    {name}({info['role']}): {info['type_code']}")

# ===== 2. 获取执行任务 =====
print("\nStep 2: 获取执行任务...")
key_execs = {
    4519: "FMS-0630", 4562: "FMS-0615", 4464: "FMS-0530",
    4651: "FMS-0715", 4665: "FMS-0730",
    4639: "海外售后v4.5-0710", 4527: "海外售后服务器迁移-0705", 4462: "海外售后v4.4-0608",
    4671: "FMS-0830", 4672: "FMS-0815",
}

all_tasks = []
for eid, ename in key_execs.items():
    resp = api_get(f"/executions/{eid}/tasks?status=all&recPerPage=1000")
    tasks = resp.get('tasks', []) if isinstance(resp, dict) else (resp if isinstance(resp, list) else [])
    for t in tasks:
        all_tasks.append({
            'id': str(t['id']), 'name': t.get('name', ''),
            'type': t.get('type', ''), 'assignedTo': t.get('assignedTo', ''),
            'executionName': ename, 'finishedDate': t.get('finishedDate', ''),
            'consumed': float(t.get('consumed', 0) or 0), 'status': t.get('status', ''),
        })

print(f"  共 {len(all_tasks)} 条任务")

# ===== 3. 匹配校验 =====
print("\nStep 3: 校验...")
filtered = [t for t in all_tasks if t['finishedDate'] and t['finishedDate'] >= '2026-07-01' and t['finishedDate'] <= '2026-07-08']

results = []
unknown_persons = set()

for t in filtered:
    account = t['assignedTo']
    person = account_to_name.get(account) or account_to_name.get(account.lower()) or account_to_name.get(account.upper())
    
    if not person:
        unknown_persons.add(account)
        continue
    
    expected = person_expected.get(person)
    if not expected:
        continue
    
    actual_type = t['type']
    task_name = t['name']
    
    if expected.get('multi_type'):
        # 多类型：按任务名称匹配期望类型
        matched_expected = None
        for pattern, tcode in PRODUCT_TASK_PATTERNS:
            if re.search(pattern, task_name):
                matched_expected = tcode
                break
        # 如果都没匹配到，用第一个（需求设计）作为默认
        if matched_expected is None:
            matched_expected = 'ab2_request_des'
        
        match = (actual_type == matched_expected)
        expected_display = matched_expected
    else:
        expected_display = expected['type_code']
        match = (actual_type == expected_display)
    
    results.append({
        '人员': person, '岗位': expected['role'],
        '禅道账号': account,
        '期望类型': expected_display,
        '实际类型': actual_type, '匹配': match,
        '任务ID': t['id'], '任务名称': task_name,
        '执行': t['executionName'], '完成日期': t['finishedDate'], '工时': t['consumed'],
        'multi_type': expected.get('multi_type', False),
    })

# ===== 4. 输出 =====
mismatch = [r for r in results if not r['匹配']]
matched = [r for r in results if r['匹配']]

print(f"\n===== 校验结果 =====")
print(f"7/1-7/8完成: {len(results)} 条, 匹配: {len(matched)}, 不匹配: {len(mismatch)}")
if unknown_persons:
    print(f"未匹配账号: {unknown_persons}")

by_person = defaultdict(lambda: {'total': 0, 'match': 0, 'mismatch': 0, 'details': []})
for r in results:
    by_person[r['人员']]['total'] += 1
    if r['匹配']:
        by_person[r['人员']]['match'] += 1
    else:
        by_person[r['人员']]['mismatch'] += 1
    by_person[r['人员']]['details'].append(r)

print(f"\n按人汇总:")
for name, stats in sorted(by_person.items()):
    exp = person_expected[name]
    if exp.get('multi_type'):
        exp_str = '多类型(按名称匹配)'
    else:
        exp_str = exp['type_code']
    status = '✅' if stats['mismatch'] == 0 else '⚠️'
    print(f"  {status} {name}({exp['role']}, {exp_str}): {stats['total']}条, 匹配{stats['match']}, 不匹配{stats['mismatch']}")
    if stats['mismatch'] > 0:
        for d in stats['details']:
            if not d['匹配']:
                print(f"      ❌ 期望={d['期望类型']} 实际={d['实际类型']} | {d['任务名称'][:60]} | {d['执行']}")

# ===== 输出JSON =====
output = {
    'summary': {
        '总任务数': len(results), '匹配数': len(matched),
        '不匹配数': len(mismatch), '未匹配账号': list(unknown_persons),
    },
    'person_summary': {
        name: {
            'total': s['total'], 'match': s['match'], 'mismatch': s['mismatch'],
            'role': person_expected[name]['role'],
            'expected_type': '多类型' if person_expected[name].get('multi_type') else person_expected[name]['type_code'],
            'multi_type': person_expected[name].get('multi_type', False),
        }
        for name, s in by_person.items()
    },
    'mismatch_details': mismatch,
    'match_details': matched,
}

with open('/Users/crystal/WorkBuddy/禅道任务/task_type_audit.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\n写入: /Users/crystal/WorkBuddy/禅道任务/task_type_audit.json")
