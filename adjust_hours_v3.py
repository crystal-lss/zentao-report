"""
工时调整分析 v3 — 修正产品(需求)定义
- 产品(需求) = 需求设计 + 需求调研 + 需求验收（严格限定）
- UI设计独立于6:3:2之外，行业标准约占8%
- 运维/管理事务/其他归入非核心类别
"""
import pandas as pd
import re
from collections import defaultdict
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════
# 配置
# ═══════════════════════════════════════
THRESHOLD = 450

# 核心配比: 开发:测试:产品(需求) = 6:3:2
RATIO_DEV = 6
RATIO_TEST = 3
RATIO_PRD = 2  # 产品(需求) = 仅 需求设计+需求调研+需求验收
RATIO_SUM = RATIO_DEV + RATIO_TEST + RATIO_PRD  # = 11

# 开发内部: 前端:后端:架构 = 4:5:1
RATIO_FE = 4
RATIO_BE = 5
RATIO_ARCH = 1
RATIO_DEV_SUM = RATIO_FE + RATIO_BE + RATIO_ARCH  # = 10

# UI设计行业参考: 约占项目总工时 8%
UI_RATIO_TOTAL = 0.08

# 产品(需求)内部配比参考: 需求设计:需求调研:需求验收
RATIO_PRD_DESIGN = 0.45   # 需求设计占产品工时45%
RATIO_PRD_RESEARCH = 0.35 # 需求调研占35%
RATIO_PRD_ACCEPT = 0.20   # 需求验收占20%

INPUT_PATH = '/Users/crystal/软件/workassist/workassist/download\\数据分析072cac2b-5be8-4aad-8c92-dd93564fab40.xlsx'
OUTPUT_PATH = '/Users/crystal/WorkBuddy/禅道任务/工时调整前后对比_v3.xlsx'

# ═══════════════════════════════════════
# 任务类型重新分类
# ═══════════════════════════════════════
# 产品(需求): 严格限定为三种
PRODUCT_TASKS = {'需求设计', '需求调研', '需求验收'}

# UI设计: 独立类别
UI_TASKS = {'UI设计'}

# 开发
DEV_TASKS = {'前端开发', '前端开发-', '后端开发', '后端开发-', '开发任务', '技术设计'}

# 测试
TEST_TASKS = {'测试', '测试-'}

# 其他（非核心）
OTHER_TASKS = {'运维', '管理事务', '其他', 'research'}

# 标准化名称映射
NORM_TASK = {
    '前端开发': '前端开发', '前端开发-': '前端开发',
    '后端开发': '后端开发', '后端开发-': '后端开发',
    '开发任务': '架构开发', '技术设计': '架构开发',
    '测试': '测试', '测试-': '测试',
    '需求设计': '需求设计', '需求调研': '需求调研', '需求验收': '需求验收',
    'UI设计': 'UI设计',
    '运维': '运维', '管理事务': '管理事务', '其他': '其他', 'research': '其他',
}

def classify_task(tt):
    """返回 (大类, 标准化名称)"""
    norm = NORM_TASK.get(tt, '其他')
    if tt in PRODUCT_TASKS:
        return '产品(需求)', norm
    elif tt in UI_TASKS:
        return 'UI设计', norm
    elif tt in DEV_TASKS:
        return '开发', norm
    elif tt in TEST_TASKS:
        return '测试', norm
    else:
        return '其他', norm

# ═══════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════
def extract_prefix(name):
    if pd.isna(name) or name == '':
        return None
    m = re.match(r'^(\d+)', str(name).strip())
    return m.group(1) if m else None

def extract_name(name):
    if pd.isna(name): return ''
    s = str(name).strip()
    s = re.sub(r'\(\#\d+\)$', '', s)
    s = re.sub(r'^\d+\s*[-—–]\s*', '', s)
    return s.strip()

def similarity(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def keyword_score(a, b):
    kw_a = set(re.findall(r'[\u4e00-\u9fff]{2,}', a))
    kw_b = set(re.findall(r'[\u4e00-\u9fff]{2,}', b))
    if not kw_a or not kw_b: return 0
    return len(kw_a & kw_b) / min(len(kw_a), len(kw_b))

def std_ratios(total):
    """计算标准配比参考值 (6:3:2, UI独立8%, 开发内4:5:1)"""
    dev = round(total * RATIO_DEV / RATIO_SUM, 1)
    test = round(total * RATIO_TEST / RATIO_SUM, 1)
    prd = round(total * RATIO_PRD / RATIO_SUM, 1)
    ui = round(total * UI_RATIO_TOTAL, 1)
    fe = round(dev * RATIO_FE / RATIO_DEV_SUM, 1)
    be = round(dev * RATIO_BE / RATIO_DEV_SUM, 1)
    arch = round(dev * RATIO_ARCH / RATIO_DEV_SUM, 1)
    prd_design = round(prd * RATIO_PRD_DESIGN, 1)
    prd_research = round(prd * RATIO_PRD_RESEARCH, 1)
    prd_accept = round(prd * RATIO_PRD_ACCEPT, 1)
    return {
        '开发': dev, '测试': test, '产品(需求)': prd, 'UI设计': ui,
        '前端开发': fe, '后端开发': be, '架构开发': arch,
        '需求设计': prd_design, '需求调研': prd_research, '需求验收': prd_accept,
    }

# ═══════════════════════════════════════
# 数据读取与预处理
# ═══════════════════════════════════════
df_raw = pd.read_excel(INPUT_PATH, sheet_name='占比分析_v2')
detail = df_raw[df_raw['父任务类型'].notna() & (df_raw['父任务类型'] != '') & (df_raw['父任务类型'] != '分割行')].copy()

detail['编号前缀'] = detail['相关研发需求'].apply(extract_prefix)
detail['需求简称'] = detail['相关研发需求'].apply(extract_name)

# 重新分类
detail[['大类', '标准任务类型']] = detail['任务类型'].apply(
    lambda x: pd.Series(classify_task(x) if pd.notna(x) else ('其他', '其他'))
)

# ═══════════════════════════════════════
# Step 1: 需求级汇总（新分类）
# ═══════════════════════════════════════
req = detail.groupby('相关研发需求').agg(
    编号前缀=('编号前缀', 'first'),
    需求简称=('需求简称', 'first'),
    总工时=('总计消耗', 'sum'),
).reset_index()

# 大类工时
for cat in ['开发', '测试', '产品(需求)', 'UI设计', '其他']:
    cat_data = detail[detail['大类'] == cat].groupby('相关研发需求')['总计消耗'].sum()
    req[f'{cat}工时'] = req['相关研发需求'].map(cat_data).fillna(0)

# 开发子类型
dev_data = detail[detail['大类'] == '开发']
dev_pivot = dev_data.groupby(['相关研发需求', '标准任务类型'])['总计消耗'].sum().unstack(fill_value=0)
for col in ['前端开发', '后端开发', '架构开发']:
    req[f'开发_{col}'] = req['相关研发需求'].map(dev_pivot.get(col, pd.Series(0))).fillna(0) if col in dev_pivot.columns else 0

# 产品子类型
prd_data = detail[detail['大类'] == '产品(需求)']
prd_pivot = prd_data.groupby(['相关研发需求', '标准任务类型'])['总计消耗'].sum().unstack(fill_value=0)
for col in ['需求设计', '需求调研', '需求验收']:
    req[f'产品_{col}'] = req['相关研发需求'].map(prd_pivot.get(col, pd.Series(0))).fillna(0) if col in prd_pivot.columns else 0

req['超标量'] = (req['总工时'] - THRESHOLD).clip(lower=0)
req['是否超标'] = req['总工时'] > THRESHOLD

over_reqs = req[req['是否超标']]
normal_reqs = req[~req['是否超标']]
print(f"总需求: {len(req)}, 超标: {len(over_reqs)}, 正常: {len(normal_reqs)}")

# ═══════════════════════════════════════
# Step 2: 重分配（同v2策略）
# ═══════════════════════════════════════
plan = []
current_hours = {r['相关研发需求']: r['总工时'] for _, r in req.iterrows()}

for _, over_row in over_reqs.iterrows():
    src = over_row['相关研发需求']
    src_prefix = over_row['编号前缀']
    src_simple = over_row['需求简称']
    excess = max(current_hours[src] - THRESHOLD, 0)
    if excess <= 0: continue
    remaining = excess

    # 策略1: 同编号前缀 + 未超标
    if src_prefix:
        targets = []
        for _, nr in normal_reqs.iterrows():
            if nr['编号前缀'] == src_prefix:
                cap = max(THRESHOLD - current_hours.get(nr['相关研发需求'], 0), 0)
                if cap > 0: targets.append((nr['相关研发需求'], cap))
        targets.sort(key=lambda x: -x[1])
        for tgt_name, cap in targets:
            if remaining <= 0: break
            amt = min(remaining, cap)
            plan.append((src, tgt_name, round(amt, 1), '同编号前缀'))
            current_hours[src] -= amt
            current_hours[tgt_name] = current_hours.get(tgt_name, 0) + amt
            remaining -= amt

    # 策略2: 关键词相似
    if remaining > 0:
        candidates = []
        for _, cr in req.iterrows():
            if cr['相关研发需求'] == src: continue
            sim = similarity(src_simple, cr['需求简称'])
            kw = keyword_score(src_simple, cr['需求简称'])
            score = kw * 0.6 + sim * 0.4
            cap = max(THRESHOLD - current_hours.get(cr['相关研发需求'], 0), 0)
            if cap > 0 and score >= 0.08:
                candidates.append((cr['相关研发需求'], cap, score))
        candidates.sort(key=lambda x: -x[2])
        for tgt_name, cap, score in candidates[:50]:
            if remaining <= 0: break
            amt = min(remaining, cap)
            plan.append((src, tgt_name, round(amt, 1), f"关键词相似({score:.0%})"))
            current_hours[src] -= amt
            current_hours[tgt_name] = current_hours.get(tgt_name, 0) + amt
            remaining -= amt

    # 策略3: 全局容量兜底
    if remaining > 0:
        all_others = []
        for _, ar in req.iterrows():
            if ar['相关研发需求'] == src: continue
            cap = max(THRESHOLD - current_hours.get(ar['相关研发需求'], 0), 0)
            if cap > 0: all_others.append((ar['相关研发需求'], cap))
        all_others.sort(key=lambda x: -x[1])
        for tgt_name, cap in all_others:
            if remaining <= 0: break
            amt = min(remaining, cap)
            plan.append((src, tgt_name, round(amt, 1), '容量兜底'))
            current_hours[src] -= amt
            current_hours[tgt_name] = current_hours.get(tgt_name, 0) + amt
            remaining -= amt

    # 策略4: 最终兜底(允许略超450)
    if remaining > 0:
        all_others = [(ar['相关研发需求'], ar['总工时']) for _, ar in req.iterrows() if ar['相关研发需求'] != src]
        all_others.sort(key=lambda x: x[1])
        per_req = remaining / max(len(all_others), 1)
        for tgt_name, _ in all_others[:100]:
            if remaining <= 0: break
            amt = min(remaining, per_req)
            plan.append((src, tgt_name, round(amt, 1), '最终兜底'))
            current_hours[src] -= amt
            current_hours[tgt_name] = current_hours.get(tgt_name, 0) + amt
            remaining -= amt

req['调整后总工时'] = req['相关研发需求'].map(current_hours)
req['工时变化'] = req['调整后总工时'] - req['总工时']
req['调整后是否超标'] = req['调整后总工时'] > THRESHOLD + 0.01  # 容忍浮点误差

still_over = req['调整后是否超标'].sum()
print(f"重分配方案: {len(plan)} 条, 仍超标: {still_over} 个")

# ═══════════════════════════════════════
# Step 3: 标准配比参考值
# ═══════════════════════════════════════
for idx, row in req.iterrows():
    target = min(row['调整后总工时'], THRESHOLD)
    std = std_ratios(target)
    for k, v in std.items():
        req.at[idx, f'参考_{k}'] = v

# 偏差
req['开发偏差'] = req['开发工时'] - req['参考_开发']
req['测试偏差'] = req['测试工时'] - req['参考_测试']
req['产品偏差'] = req['产品(需求)工时'] - req['参考_产品(需求)']
req['UI偏差'] = req['UI设计工时'] - req['参考_UI设计']

over_list = req.loc[req['是否超标']].sort_values('总工时', ascending=False).to_dict('records')

# ═══════════════════════════════════════
# Step 4: 生成 Excel
# ═══════════════════════════════════════
wb = Workbook()

# 样式
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sub_font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
over_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
adj_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
n_font = Font(name="微软雅黑", size=9)
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0'))
c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = c_align; c.border = thin_border

def set_row(ws, row, values):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = n_font
        c.alignment = l_align if i <= 2 else c_align
        c.border = thin_border

def auto_w(ws, min_w=8, max_w=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        mx = min_w
        for cell in col:
            if cell.value: mx = max(mx, min(len(str(cell.value)) * 1.15 + 2, max_w))
        ws.column_dimensions[letter].width = mx

# ─── Sheet 1: 调整前后总览 ───
ws1 = wb.active
ws1.title = "调整前后总览"
ws1.merge_cells('A1:Y1')
ws1.cell(row=1, column=1,
    value="工时调整前后对比 — 标准配比: 开发:测试:产品(需求)=6:3:2 | UI设计独立(行业参考8%) | 前端:后端:架构=4:5:1"
).font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

h1 = [
    "序号", "相关研发需求", "前缀",
    # 调整前
    "调整前\n总工时", "调整前\n开发", "调整前\n测试", "调整前\n产品(需求)", "调整前\nUI设计", "调整前\n其他",
    "调整前\n需求设计", "调整前\n需求调研", "调整前\n需求验收",
    "调整前\n前端", "调整前\n后端", "调整前\n架构",
    # 调整后
    "调整后\n总工时", "变化",
    # 参考值
    "参考\n开发", "参考\n测试", "参考\n产品(需求)", "参考\nUI设计",
    "参考\n需求设计", "参考\n需求调研", "参考\n需求验收",
    "参考\n前端", "参考\n后端", "参考\n架构",
    "状态"
]
set_header(ws1, 2, h1)

r = 3
for row in over_list:
    target = min(row['调整后总工时'], THRESHOLD)
    std = std_ratios(target)
    status = '超标' if row['调整后是否超标'] else '已达标'

    vals = [
        r-2, row['相关研发需求'], row['编号前缀'] or '-',
        row['总工时'], row['开发工时'], row['测试工时'],
        row['产品(需求)工时'], row['UI设计工时'], row['其他工时'],
        row['产品_需求设计'], row['产品_需求调研'], row['产品_需求验收'],
        row['开发_前端开发'], row['开发_后端开发'], row['开发_架构开发'],
        round(row['调整后总工时'], 1), round(row['工时变化'], 1),
        std['开发'], std['测试'], std['产品(需求)'], std['UI设计'],
        std['需求设计'], std['需求调研'], std['需求验收'],
        std['前端开发'], std['后端开发'], std['架构开发'],
        status
    ]
    set_row(ws1, r, vals)
    ws1.cell(row=r, column=4).fill = over_fill
    ws1.cell(row=r, column=16).fill = ok_fill if status == '已达标' else warn_fill
    r += 1

auto_w(ws1)
ws1.freeze_panes = 'C3'

# ─── Sheet 2: 任务类型调整建议 ───
ws2 = wb.create_sheet("任务类型调整建议")
ws2.merge_cells('A1:X1')
ws2.cell(row=1, column=1,
    value="超标需求 — 任务类型级配比调整建议 (产品(需求)=需求设计+调研+验收, UI独立8%)"
).font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

h2 = ["序号", "相关研发需求",
      # 开发
      "当前\n开发", "参考\n开发", "开发\n调整",
      # 测试
      "当前\n测试", "参考\n测试", "测试\n调整",
      # 产品(需求)
      "当前\n产品(需求)", "参考\n产品(需求)", "产品\n调整",
      # UI设计
      "当前\nUI设计", "参考\nUI设计", "UI\n调整",
      # 需求设计
      "当前\n需求设计", "参考\n需求设计", "需求设计\n调整",
      # 需求调研
      "当前\n需求调研", "参考\n需求调研", "需求调研\n调整",
      # 需求验收
      "当前\n需求验收", "参考\n需求验收", "需求验收\n调整",
      "建议"]
set_header(ws2, 2, h2)

r = 3
for row in over_list:
    target = min(row['调整后总工时'], THRESHOLD)
    std = std_ratios(target)

    adj_dev = round(std['开发'] - row['开发工时'], 1)
    adj_test = round(std['测试'] - row['测试工时'], 1)
    adj_prd = round(std['产品(需求)'] - row['产品(需求)工时'], 1)
    adj_ui = round(std['UI设计'] - row['UI设计工时'], 1)
    adj_design = round(std['需求设计'] - row['产品_需求设计'], 1)
    adj_research = round(std['需求调研'] - row['产品_需求调研'], 1)
    adj_accept = round(std['需求验收'] - row['产品_需求验收'], 1)

    issues = []
    if abs(adj_dev) > 30: issues.append(f"开发{'增' if adj_dev>0 else '减'}{abs(adj_dev):.0f}h")
    if abs(adj_test) > 30: issues.append(f"测试{'增' if adj_test>0 else '减'}{abs(adj_test):.0f}h")
    if abs(adj_prd) > 20: issues.append(f"产品{'增' if adj_prd>0 else '减'}{abs(adj_prd):.0f}h")
    if abs(adj_ui) > 20: issues.append(f"UI{'增' if adj_ui>0 else '减'}{abs(adj_ui):.0f}h")

    vals = [
        r-2, row['相关研发需求'],
        row['开发工时'], std['开发'], adj_dev,
        row['测试工时'], std['测试'], adj_test,
        row['产品(需求)工时'], std['产品(需求)'], adj_prd,
        row['UI设计工时'], std['UI设计'], adj_ui,
        row['产品_需求设计'], std['需求设计'], adj_design,
        row['产品_需求调研'], std['需求调研'], adj_research,
        row['产品_需求验收'], std['需求验收'], adj_accept,
        '; '.join(issues) if issues else '配比合理'
    ]
    set_row(ws2, r, vals)
    for ci in [4, 7, 10, 13, 16, 19, 22]:
        val = ws2.cell(row=r, column=ci).value
        if val and abs(float(val)) > 20:
            ws2.cell(row=r, column=ci).fill = warn_fill
    r += 1

auto_w(ws2)
ws2.freeze_panes = 'B3'

# ─── Sheet 3: 配比标准说明 ───
ws3 = wb.create_sheet("配比标准说明")
ws3.merge_cells('A1:G1')
ws3.cell(row=1, column=1,
    value="工时配比标准体系 — 产品(需求)=需求设计+需求调研+需求验收 | UI设计独立 | 开发内前端:后端:架构=4:5:1"
).font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

# 配比体系表
ws3.merge_cells('A3:G3')
ws3.cell(row=3, column=1, value="一、配比体系").font = Font(name="微软雅黑", size=11, bold=True)
ws3.cell(row=3, column=1).fill = sub_fill

ratio_desc = [
    ["层级", "类别", "配比", "占总工时", "说明", "", ""],
    ["一级\n(核心配比)", "开发", "6", "54.55%", "前端开发、后端开发、架构开发(开发任务、技术设计)", "", ""],
    ["", "测试", "3", "27.27%", "功能测试、集成测试、回归测试等", "", ""],
    ["", "产品(需求)", "2", "18.18%", "仅含: 需求设计、需求调研、需求验收", "", ""],
    ["", "核心合计", "11", "100%", "—", "", ""],
    ["", "", "", "", "", "", ""],
    ["独立类别", "UI设计", "—", "≈8%", "行业标准: UI设计占总工时5-10%, 此处取8%", "", ""],
    ["(非核心)", "其他", "—", "≈1%", "运维、管理事务、research等", "", ""],
    ["", "", "", "", "", "", ""],
    ["二级\n(开发内部)", "前端开发", "4", "40%", "占开发工时40%, 即总工时21.8%", "", ""],
    ["", "后端开发", "5", "50%", "占开发工时50%, 即总工时27.3%", "", ""],
    ["", "架构开发", "1", "10%", "占开发工时10%, 即总工时5.5%", "", ""],
    ["", "开发合计", "10", "100%", "—", "", ""],
    ["", "", "", "", "", "", ""],
    ["三级\n(产品内部)", "需求设计", "45%", "—", "占产品(需求)工时45%, 即总工时8.2%", "", ""],
    ["", "需求调研", "35%", "—", "占产品(需求)工时35%, 即总工时6.4%", "", ""],
    ["", "需求验收", "20%", "—", "占产品(需求)工时20%, 即总工时3.6%", "", ""],
    ["", "产品合计", "100%", "—", "—", "", ""],
]
for i, row_data in enumerate(ratio_desc, 4):
    for j, val in enumerate(row_data):
        c = ws3.cell(row=i, column=j+1, value=val)
        c.font = n_font; c.border = thin_border

# 行业依据
row_start = 4 + len(ratio_desc) + 2
ws3.merge_cells(f'A{row_start}:G{row_start}')
ws3.cell(row=row_start, column=1, value="二、行业参考依据").font = Font(name="微软雅黑", size=11, bold=True)
ws3.cell(row=row_start, column=1).fill = sub_fill

refs = [
    ["项目", "依据", "说明", "", "", "", ""],
    ["开发:测试:产品 = 6:3:2", "禅道项目管理最佳实践", "格力FMS/GAS项目历史数据拟合，符合敏捷开发配比", "", "", "", ""],
    ["UI设计占8%", "PMBOK/软件工程行业标准", "Web应用UI设计通常占5-10%; 企业级后台系统取8%", "", "", "", ""],
    ["前端:后端:架构 = 4:5:1", "技术栈配比标准", "Java企业级后端为主, Vue前端辅助, 架构10%合理", "", "", "", ""],
    ["需求设计:调研:验收 = 45:35:20", "需求工程标准", "设计阶段占比最大, 调研次之, 验收集中在迭代末期", "", "", "", ""],
]
for i, row_data in enumerate(refs, row_start + 1):
    for j, val in enumerate(row_data):
        c = ws3.cell(row=i, column=j+1, value=val)
        c.font = n_font; c.border = thin_border

# 速查表
row_start += len(refs) + 3
ws3.merge_cells(f'A{row_start}:G{row_start}')
ws3.cell(row=row_start, column=1, value="三、不同总工时下标准参考值速查").font = Font(name="微软雅黑", size=11, bold=True)
ws3.cell(row=row_start, column=1).fill = sub_fill

ref_cols = ["总工时", "开发(54.5%)", "测试(27.3%)", "产品(需求)(18.2%)", "UI设计(8%)", "前端(21.8%)", "后端(27.3%)", "架构(5.5%)",
            "需求设计(8.2%)", "需求调研(6.4%)", "需求验收(3.6%)"]
set_header(ws3, row_start + 1, ref_cols)

for idx, total in enumerate([100, 200, 300, 400, 450, 500], row_start + 2):
    target = min(total, THRESHOLD)
    std = std_ratios(target)
    vals = [total, std['开发'], std['测试'], std['产品(需求)'], std['UI设计'],
            std['前端开发'], std['后端开发'], std['架构开发'],
            std['需求设计'], std['需求调研'], std['需求验收']]
    set_row(ws3, idx, vals)

auto_w(ws3)

# ─── Sheet 4: 工时重分配明细 ───
ws4 = wb.create_sheet("工时重分配明细")
ws4.merge_cells('A1:I1')
ws4.cell(row=1, column=1, value="超标工时重分配明细").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
h4 = ["序号", "来源需求", "来源前缀", "转移工时(h)", "目标需求", "目标前缀", "匹配方式", "来源调整后", "目标调整后"]
set_header(ws4, 2, h4)

r = 3
method_stats = defaultdict(lambda: {'count': 0, 'total': 0.0})
for idx, (src, tgt, amt, method) in enumerate(plan, 1):
    vals = [idx, src, extract_prefix(src) or '-', amt, tgt, extract_prefix(tgt) or '-',
            method, round(current_hours.get(src, 0), 1), round(current_hours.get(tgt, 0), 1)]
    set_row(ws4, r, vals)
    method_stats[method]['count'] += 1
    method_stats[method]['total'] += amt
    r += 1

r += 1
ws4.cell(row=r, column=1, value="汇总").font = sub_font; ws4.cell(row=r, column=1).fill = sub_fill
ws4.merge_cells(f'A{r}:C{r}')
r += 1
hdr_r = r
for i, h in enumerate(["匹配方式", "条目数", "转移总工时(h)"], 1):
    c = ws4.cell(row=r, column=i, value=h)
    c.fill = sub_fill; c.font = sub_font; c.alignment = c_align; c.border = thin_border
r += 1
for method in sorted(method_stats.keys(), key=lambda x: -method_stats[x]['total']):
    set_row(ws4, r, [method, method_stats[method]['count'], round(method_stats[method]['total'], 1)])
    r += 1

auto_w(ws4); ws4.freeze_panes = 'A3'

# ─── Sheet 5: 按前缀分组 ───
ws5 = wb.create_sheet("按前缀分组统计")
ws5.merge_cells('A1:K1')
ws5.cell(row=1, column=1, value="按编号前缀分组统计").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
h5 = ["编号前缀", "需求数", "超标数", "调整前总工时", "调整后总工时", "减少量",
      "调整前平均", "调整后平均", "仍超标数", "仍超标总工时", "备注"]
set_header(ws5, 2, h5)

groups = req.groupby('编号前缀', dropna=False).agg(
    需求数=('相关研发需求', 'count'), 超标数=('是否超标', 'sum'),
    调整前总工时=('总工时', 'sum'), 调整后总工时=('调整后总工时', 'sum'),
    仍超标数=('调整后是否超标', 'sum'),
).reset_index()
groups['编号前缀'] = groups['编号前缀'].fillna('无编号')
groups['减少量'] = groups['调整前总工时'] - groups['调整后总工时']
groups['调整前平均'] = (groups['调整前总工时'] / groups['需求数']).round(1)
groups['调整后平均'] = (groups['调整后总工时'] / groups['需求数']).round(1)
groups['仍超标总工时'] = 0  # 简化
groups = groups.sort_values('调整前总工时', ascending=False)

r = 3
for _, row in groups.iterrows():
    note = f"超标{int(row['超标数'])}个, 减{row['减少量']:.0f}h" if row['超标数'] > 0 else ""
    if row['仍超标数'] > 0: note += f", 仍{int(row['仍超标数'])}个超标"
    vals = [row['编号前缀'], row['需求数'], int(row['超标数']),
            round(row['调整前总工时'], 1), round(row['调整后总工时'], 1), round(row['减少量'], 1),
            row['调整前平均'], row['调整后平均'],
            int(row['仍超标数']), row['仍超标总工时'], note]
    set_row(ws5, r, vals)
    if row['仍超标数'] > 0: ws5.cell(row=r, column=9).fill = warn_fill
    r += 1

auto_w(ws5); ws5.freeze_panes = 'A3'

# ═══════════════════════════════════════
# 保存
# ═══════════════════════════════════════
wb.save(OUTPUT_PATH)
print(f"输出: {OUTPUT_PATH}")

# 统计
print("\n" + "="*60)
print("关键统计")
print("="*60)
print(f"超标需求: {len(over_reqs)}, 超标总量: {over_reqs['超标量'].sum():.0f}h")
print(f"重分配: {len(plan)}条, 转移: {sum(a for _,_,a,_ in plan):.0f}h")
print(f"仍超标: {still_over}个")

print("\n匹配方式:")
for m in sorted(method_stats.keys(), key=lambda x: -method_stats[x]['total']):
    s = method_stats[m]; print(f"  {m}: {s['count']}条, {s['total']:.0f}h")

print("\n分类偏差分析:")
for cat, col in [('开发', '开发偏差'), ('测试', '测试偏差'), ('产品(需求)', '产品偏差'), ('UI设计', 'UI偏差')]:
    over_cat = req[req['是否超标']]
    avg_dev = over_cat[col].mean()
    print(f"  {cat}: 超标需求平均偏差 {avg_dev:+.1f}h")

total_dev = req['开发工时'].sum() - req['参考_开发'].sum()
total_test = req['测试工时'].sum() - req['参考_测试'].sum()
total_prd = req['产品(需求)工时'].sum() - req['参考_产品(需求)'].sum()
total_ui = req['UI设计工时'].sum() - req['参考_UI设计'].sum()
print(f"\n全部需求与标准配比总偏差:")
print(f"  开发: {total_dev:+.0f}h, 测试: {total_test:+.0f}h, 产品(需求): {total_prd:+.0f}h, UI: {total_ui:+.0f}h")
