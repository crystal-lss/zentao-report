"""
工时调整分析脚本 v2 — 增强版
- 需求工时 > 450h 的全量重分配（含智能相似度匹配）
- 标准配比 开发:测试:产品 = 6:3:2
- 开发细分 前端:后端:架构 = 4:5:1
- 任务类型级调整前后对比
"""
import pandas as pd
import re
import numpy as np
from collections import defaultdict
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════
# 配置
# ════════════════════════════════════════════════════════
THRESHOLD = 450
RATIO_DEV_TEST_PRD = {'开发': 6, '测试': 3, '产品': 2}  # 6:3:2
RATIO_FE_BE_ARCH = {'前端开发': 4, '后端开发': 5, '架构开发': 1}  # 4:5:1
UI_RATIO = 0.15  # UI 占产品工时 15%

INPUT_PATH = '/Users/crystal/软件/workassist/workassist/download\\数据分析072cac2b-5be8-4aad-8c92-dd93564fab40.xlsx'
OUTPUT_PATH = '/Users/crystal/WorkBuddy/禅道任务/工时调整前后对比_v2.xlsx'

# ════════════════════════════════════════════════════════
# 工具函数
# ════════════════════════════════════════════════════════
def extract_prefix(name):
    if pd.isna(name) or name == '':
        return None
    m = re.match(r'^(\d+)', str(name).strip())
    return m.group(1) if m else None

def extract_name(name):
    if pd.isna(name):
        return ''
    s = str(name).strip()
    s = re.sub(r'\(\#\d+\)$', '', s)
    s = re.sub(r'^\d+\s*[-—–]\s*', '', s)
    return s.strip()

def similarity(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def keyword_score(a, b):
    """基于关键词重叠的评分"""
    kw_a = set(re.findall(r'[\u4e00-\u9fff]{2,}', a))
    kw_b = set(re.findall(r'[\u4e00-\u9fff]{2,}', b))
    if not kw_a or not kw_b:
        return 0
    intersection = kw_a & kw_b
    return len(intersection) / min(len(kw_a), len(kw_b))

TASK_TYPE_MAP = {
    '前端开发': '前端开发', '前端开发-': '前端开发',
    '后端开发': '后端开发', '后端开发-': '后端开发',
    '开发任务': '架构开发', '技术设计': '架构开发',
    '测试': '测试', '测试-': '测试',
    '需求设计': '产品', '需求调研': '产品', '需求验收': '产品',
    'UI设计': 'UI设计', '运维': '产品', '管理事务': '产品',
    '其他': '产品', 'research': '产品',
}

def calc_standard_ratios(total_hours):
    """标准配比计算"""
    ratio_sum = sum(RATIO_DEV_TEST_PRD.values())
    fe_ratio_sum = sum(RATIO_FE_BE_ARCH.values())
    dev_h = total_hours * RATIO_DEV_TEST_PRD['开发'] / ratio_sum
    test_h = total_hours * RATIO_DEV_TEST_PRD['测试'] / ratio_sum
    prd_h = total_hours * RATIO_DEV_TEST_PRD['产品'] / ratio_sum
    return {
        '开发': round(dev_h, 1),
        '测试': round(test_h, 1),
        '产品': round(prd_h, 1),
        'UI': round(prd_h * UI_RATIO, 1),
        '前端开发': round(dev_h * RATIO_FE_BE_ARCH['前端开发'] / fe_ratio_sum, 1),
        '后端开发': round(dev_h * RATIO_FE_BE_ARCH['后端开发'] / fe_ratio_sum, 1),
        '架构开发': round(dev_h * RATIO_FE_BE_ARCH['架构开发'] / fe_ratio_sum, 1),
    }

# ════════════════════════════════════════════════════════
# 数据读取
# ════════════════════════════════════════════════════════
df_raw = pd.read_excel(INPUT_PATH, sheet_name='占比分析_v2')
detail = df_raw[df_raw['父任务类型'].notna() & (df_raw['父任务类型'] != '') & (df_raw['父任务类型'] != '分割行')].copy()
detail['编号前缀'] = detail['相关研发需求'].apply(extract_prefix)
detail['需求简称'] = detail['相关研发需求'].apply(extract_name)
detail['标准化父类型'] = detail['父任务类型'].map(lambda x: x if x in ('开发','测试','产品','UI') else '未知')
detail['标准化任务类型'] = detail['任务类型'].map(TASK_TYPE_MAP).fillna('其他')

# ════════════════════════════════════════════════════════
# Step 1: 需求级汇总
# ════════════════════════════════════════════════════════
req = detail.groupby('相关研发需求').agg(
    编号前缀=('编号前缀', 'first'),
    需求简称=('需求简称', 'first'),
    总工时=('总计消耗', 'sum'),
).reset_index()

# 各父类型工时
for pt in ['开发', '测试', '产品', 'UI']:
    pt_data = detail[detail['标准化父类型'] == pt].groupby('相关研发需求')['总计消耗'].sum()
    req[f'{pt}工时'] = req['相关研发需求'].map(pt_data).fillna(0)

# 开发子类型工时
dev_data = detail[detail['标准化父类型'] == '开发'].copy()
dev_pivot = dev_data.groupby(['相关研发需求', '标准化任务类型'])['总计消耗'].sum().unstack(fill_value=0)
for col in ['前端开发', '后端开发', '架构开发']:
    req[f'开发_{col}'] = req['相关研发需求'].map(dev_pivot.get(col, pd.Series(0))).fillna(0) if col in dev_pivot.columns else 0

req['超标量'] = (req['总工时'] - THRESHOLD).clip(lower=0)
req['是否超标'] = req['总工时'] > THRESHOLD
over_reqs = req[req['是否超标']].copy()
normal_reqs = req[~req['是否超标']].copy()

print(f"总需求: {len(req)}, 超标: {len(over_reqs)}, 正常: {len(normal_reqs)}")

# ════════════════════════════════════════════════════════
# Step 2: 增强版重分配 — 多级匹配策略
# ════════════════════════════════════════════════════════
plan = []  # [(来源, 目标, 转移工时, 匹配方式)]

# 动态跟踪每个需求的"当前"工时（随着转移逐步更新）
current_hours = {r['相关研发需求']: r['总工时'] for _, r in req.iterrows()}

for _, over_row in over_reqs.iterrows():
    src = over_row['相关研发需求']
    src_prefix = over_row['编号前缀']
    src_simple = over_row['需求简称']
    excess = max(current_hours[src] - THRESHOLD, 0)
    if excess <= 0:
        continue
    
    remaining = excess
    
    # 策略1: 同编号前缀 + 未超标（最优先）
    if src_prefix:
        targets = []
        for _, nr in normal_reqs.iterrows():
            if nr['编号前缀'] == src_prefix:
                cap = max(THRESHOLD - current_hours.get(nr['相关研发需求'], 0), 0)
                if cap > 0:
                    targets.append((nr['相关研发需求'], cap))
        targets.sort(key=lambda x: -x[1])
        for tgt_name, cap in targets:
            if remaining <= 0:
                break
            amt = min(remaining, cap)
            plan.append((src, tgt_name, round(amt, 1), '同编号前缀'))
            current_hours[src] -= amt
            current_hours[tgt_name] = current_hours.get(tgt_name, 0) + amt
            remaining -= amt
    
    # 策略2: 关键词相似 + 任意需求（增强版，不做同编号分担—那会雪上加霜）
    if remaining > 0:
        candidates = []
        for _, cr in req.iterrows():
            if cr['相关研发需求'] == src:
                continue
            sim = similarity(src_simple, cr['需求简称'])
            kw = keyword_score(src_simple, cr['需求简称'])
            score = kw * 0.6 + sim * 0.4
            cap = max(THRESHOLD - current_hours.get(cr['相关研发需求'], 0), 0)
            if cap > 0 and score >= 0.08:
                candidates.append((cr['相关研发需求'], cap, score))
        candidates.sort(key=lambda x: -x[2])
        
        for tgt_name, cap, score in candidates[:50]:
            if remaining <= 0:
                break
            amt = min(remaining, cap)
            plan.append((src, tgt_name, round(amt, 1), f"关键词相似({score:.0%})"))
            current_hours[src] -= amt
            current_hours[tgt_name] = current_hours.get(tgt_name, 0) + amt
            remaining -= amt
    
    # 策略3: 全局容量兜底 — 扩大到所有有容量的需求
    if remaining > 0:
        all_others = []
        for _, ar in req.iterrows():
            if ar['相关研发需求'] == src:
                continue
            cap = max(THRESHOLD - current_hours.get(ar['相关研发需求'], 0), 0)
            if cap > 0:
                all_others.append((ar['相关研发需求'], cap))
        all_others.sort(key=lambda x: -x[1])  # 容量大的优先
        
        for tgt_name, cap in all_others:
            if remaining <= 0:
                break
            amt = min(remaining, cap)
            plan.append((src, tgt_name, round(amt, 1), '容量兜底'))
            current_hours[src] -= amt
            current_hours[tgt_name] = current_hours.get(tgt_name, 0) + amt
            remaining -= amt
    
    # 策略4: 实在兜不住，允许目标略超450（最终兜底）
    if remaining > 0:
        all_others = [(ar['相关研发需求'], ar['总工时']) for _, ar in req.iterrows() if ar['相关研发需求'] != src]
        all_others.sort(key=lambda x: x[1])  # 小需求优先
        per_req = remaining / max(len(all_others), 1)
        for tgt_name, _ in all_others[:100]:
            if remaining <= 0:
                break
            amt = min(remaining, per_req)
            plan.append((src, tgt_name, round(amt, 1), '最终兜底(超450)'))
            current_hours[src] -= amt
            current_hours[tgt_name] = current_hours.get(tgt_name, 0) + amt
            remaining -= amt

# 应用最终调整
req['调整后总工时'] = req['相关研发需求'].map(current_hours)
req['工时变化'] = req['调整后总工时'] - req['总工时']
req['调整后超标量'] = (req['调整后总工时'] - THRESHOLD).clip(lower=0)
req['调整后是否超标'] = req['调整后总工时'] > THRESHOLD

still_over = req['调整后是否超标'].sum()
print(f"重分配方案: {len(plan)} 条")
print(f"调整后仍超标: {still_over} 个")

# ════════════════════════════════════════════════════════
# Step 3: 为每个超标需求计算标准配比参考值（任务类型级）
# ════════════════════════════════════════════════════════
for idx, row in req.iterrows():
    target = min(row['调整后总工时'], THRESHOLD)
    std = calc_standard_ratios(target)
    for k, v in std.items():
        req.at[idx, f'参考_{k}'] = v

# 计算偏差
req['开发偏差'] = req['开发工时'] - req['参考_开发']
req['测试偏差'] = req['测试工时'] - req['参考_测试']
req['产品偏差'] = req['产品工时'] - req['参考_产品']

# 重新获取超标需求（从已更新的 req 中）
over_sorted = req.loc[req['是否超标']].sort_values('总工时', ascending=False)
# 转为 dict 列表避免 pandas 索引问题
over_list = over_sorted.to_dict('records')

# 为超标需求计算任务类型级别的调整建议
def calc_task_adjustment(req_row):
    """计算单一需求的任务类型调整建议"""
    target_total = min(req_row['调整后总工时'], THRESHOLD)
    std = calc_standard_ratios(target_total)
    
    adjustments = {}
    # 开发调整
    adjustments['开发'] = round(std['开发'] - req_row['开发工时'], 1)
    adjustments['测试'] = round(std['测试'] - req_row['测试工时'], 1)
    adjustments['产品'] = round(std['产品'] - req_row['产品工时'], 1)
    adjustments['UI'] = round(std['UI'] - req_row['UI工时'], 1)
    # 开发细分
    adjustments['前端开发'] = round(std['前端开发'] - req_row['开发_前端开发'], 1)
    adjustments['后端开发'] = round(std['后端开发'] - req_row['开发_后端开发'], 1)
    adjustments['架构开发'] = round(std['架构开发'] - req_row['开发_架构开发'], 1)
    return adjustments, std

# ════════════════════════════════════════════════════════
# Step 4: 生成 Excel 报告
# ════════════════════════════════════════════════════════
wb = Workbook()

# 样式
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sub_font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
over_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
adj_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
n_font = Font(name="微软雅黑", size=9)
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0'),
)
c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_header(ws, row, headers, fill=None, font=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill or hdr_fill
        c.font = font or hdr_font
        c.alignment = c_align
        c.border = thin_border

def set_row(ws, row, values):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = n_font
        c.alignment = l_align if i <= 2 else c_align
        c.border = thin_border

def auto_w(ws, min_w=8, max_w=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) * 1.15 + 2, max_w))
        ws.column_dimensions[letter].width = mx

# ──────────────────────────────────────────────────────
# Sheet 1: 需求级调整前后总览
# ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "调整前后总览"
ws1.merge_cells('A1:W1')
ws1.cell(row=1, column=1, value="工时调整前后对比 — 阈值450h | 标准配比 开发:测试:产品=6:3:2 | 前端:后端:架构=4:5:1").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

h1 = ["序号", "相关研发需求", "前缀", 
      "调整前\n总工时", "调整前\n开发", "调整前\n测试", "调整前\n产品", "调整前\nUI",
      "调整前\n前端", "调整前\n后端", "调整前\n架构",
      "调整后\n总工时", "变化量",
      "参考\n开发", "参考\n测试", "参考\n产品", "参考\nUI",
      "参考\n前端", "参考\n后端", "参考\n架构",
      "开发偏差", "测试偏差", "产品偏差", "状态"]
set_header(ws1, 2, h1)

r = 3
for row in over_list:
    adjustments, _ = calc_task_adjustment(row)
    vals = [
        r-2, row['相关研发需求'], row['编号前缀'] or '-',
        row['总工时'], row['开发工时'], row['测试工时'], row['产品工时'], row['UI工时'],
        row['开发_前端开发'], row['开发_后端开发'], row['开发_架构开发'],
        round(row['调整后总工时'], 1), round(row['工时变化'], 1),
        row['参考_开发'], row['参考_测试'], row['参考_产品'], row['参考_UI'],
        row['参考_前端开发'], row['参考_后端开发'], row['参考_架构开发'],
        round(row['开发偏差'], 1), round(row['测试偏差'], 1), round(row['产品偏差'], 1),
        '超标' if row['调整后是否超标'] else ('已达标' if row['调整后超标量'] == 0 else '部分调整')
    ]
    set_row(ws1, r, vals)
    # 颜色标记
    if row['调整后是否超标']:
        ws1.cell(row=r, column=12).fill = over_fill
    else:
        ws1.cell(row=r, column=12).fill = ok_fill
    ws1.cell(row=r, column=4).fill = over_fill  # 调整前总工时标红
    r += 1

auto_w(ws1)
ws1.freeze_panes = 'C3'

# ──────────────────────────────────────────────────────
# Sheet 2: 任务类型级调整建议（超标需求）
# ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("任务类型调整建议")
ws2.merge_cells('A1:U1')
ws2.cell(row=1, column=1, value="超标需求 → 任务类型级配比调整建议 (调整前 vs 标准参考)").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

h2 = ["序号", "相关研发需求",
      "当前\n开发工时", "参考\n开发", "开发\n调整量",
      "当前\n测试工时", "参考\n测试", "测试\n调整量",
      "当前\n产品工时", "参考\n产品", "产品\n调整量",
      "当前\nUI工时", "参考\nUI", "UI\n调整量",
      "当前\n前端", "参考\n前端", "前端\n调整量",
      "当前\n后端", "参考\n后端", "后端\n调整量",
      "当前\n架构", "参考\n架构", "架构\n调整量",
      "建议"]
set_header(ws2, 2, h2)

r = 3
for row in over_list:
    adjustments, std = calc_task_adjustment(row)
    
    # 判断主要问题
    issues = []
    if abs(adjustments['开发']) > 50:
        direction = '增加' if adjustments['开发'] > 0 else '减少'
        issues.append(f"开发{direction}{abs(adjustments['开发']):.0f}h")
    if abs(adjustments['测试']) > 30:
        direction = '增加' if adjustments['测试'] > 0 else '减少'
        issues.append(f"测试{direction}{abs(adjustments['测试']):.0f}h")
    if abs(adjustments['产品']) > 30:
        direction = '增加' if adjustments['产品'] > 0 else '减少'
        issues.append(f"产品{direction}{abs(adjustments['产品']):.0f}h")
    
    vals = [
        r-2, row['相关研发需求'],
        row['开发工时'], std['开发'], adjustments['开发'],
        row['测试工时'], std['测试'], adjustments['测试'],
        row['产品工时'], std['产品'], adjustments['产品'],
        row['UI工时'], std['UI'], adjustments['UI'],
        row['开发_前端开发'], std['前端开发'], adjustments['前端开发'],
        row['开发_后端开发'], std['后端开发'], adjustments['后端开发'],
        row['开发_架构开发'], std['架构开发'], adjustments['架构开发'],
        '; '.join(issues) if issues else '配比已合理'
    ]
    set_row(ws2, r, vals)
    # 标记需调整的列
    for col_idx in [4, 7, 10, 13, 16, 19, 22]:
        diff_val = ws2.cell(row=r, column=col_idx).value
        if diff_val and abs(float(diff_val)) > 20:
            ws2.cell(row=r, column=col_idx).fill = warn_fill
    r += 1

auto_w(ws2)
ws2.freeze_panes = 'B3'

# ──────────────────────────────────────────────────────
# Sheet 3: 工时重分配明细
# ──────────────────────────────────────────────────────
ws3 = wb.create_sheet("工时重分配明细")
ws3.merge_cells('A1:I1')
ws3.cell(row=1, column=1, value="超标工时重分配明细 — 来源需求 → 目标需求").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

h3 = ["序号", "来源需求", "来源前缀", "转移工时(h)", "目标需求", "目标前缀", "匹配方式", "来源调整后", "目标调整后"]
set_header(ws3, 2, h3)

r = 3
method_stats = defaultdict(lambda: {'count': 0, 'total': 0.0})
for idx, (src, tgt, amt, method) in enumerate(plan, 1):
    vals = [idx, src, extract_prefix(src) or '-', amt, tgt, extract_prefix(tgt) or '-', method,
            round(current_hours.get(src, 0), 1), round(current_hours.get(tgt, 0), 1)]
    set_row(ws3, r, vals)
    method_stats[method]['count'] += 1
    method_stats[method]['total'] += amt
    r += 1

# 汇总
r += 1
ws3.cell(row=r, column=1, value="匹配方式汇总").font = Font(name="微软雅黑", size=10, bold=True)
ws3.cell(row=r, column=1).fill = sub_fill
ws3.merge_cells(f'A{r}:C{r}')
r += 1
set_header(ws3, r, ["匹配方式", "条目数", "转移总工时(h)"], sub_fill, sub_font)
r += 1
for method in sorted(method_stats.keys()):
    vals = [method, method_stats[method]['count'], round(method_stats[method]['total'], 1)]
    set_row(ws3, r, vals)
    r += 1

auto_w(ws3)
ws3.freeze_panes = 'A3'

# ──────────────────────────────────────────────────────
# Sheet 4: 按前缀分组统计
# ──────────────────────────────────────────────────────
ws4 = wb.create_sheet("按前缀分组统计")
ws4.merge_cells('A1:K1')
ws4.cell(row=1, column=1, value="按编号前缀分组 — 工时分布与调整效果").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

h4 = ["编号前缀", "需求数", "超标数", "调整前总工时", "调整后总工时", "减少量",
      "调整前平均", "调整后平均", "仍超标数", "仍超标总工时", "备注"]
set_header(ws4, 2, h4)

groups = req.groupby('编号前缀', dropna=False).agg(
    需求数=('相关研发需求', 'count'),
    超标数=('是否超标', 'sum'),
    调整前总工时=('总工时', 'sum'),
    调整后总工时=('调整后总工时', 'sum'),
    仍超标数=('调整后是否超标', 'sum'),
    仍超标总工时=('调整后超标量', 'sum'),
).reset_index()
groups['编号前缀'] = groups['编号前缀'].fillna('无编号')
groups['减少量'] = groups['调整前总工时'] - groups['调整后总工时']
groups['调整前平均'] = (groups['调整前总工时'] / groups['需求数']).round(1)
groups['调整后平均'] = (groups['调整后总工时'] / groups['需求数']).round(1)
groups = groups.sort_values('调整前总工时', ascending=False)

r = 3
for _, row in groups.iterrows():
    note = ""
    if row['超标数'] > 0:
        note = f"超标{row['超标数']}个, 共减{row['减少量']:.0f}h"
    if row['仍超标数'] > 0:
        note += f", 仍{row['仍超标数']}个超标"
    vals = [row['编号前缀'], row['需求数'], row['超标数'],
            round(row['调整前总工时'], 1), round(row['调整后总工时'], 1), round(row['减少量'], 1),
            row['调整前平均'], row['调整后平均'],
            row['仍超标数'], round(row['仍超标总工时'], 1), note]
    set_row(ws4, r, vals)
    if row['仍超标数'] > 0:
        ws4.cell(row=r, column=9).fill = warn_fill
    r += 1

auto_w(ws4)
ws4.freeze_panes = 'A3'

# ──────────────────────────────────────────────────────
# Sheet 5: 标准配比参考表
# ──────────────────────────────────────────────────────
ws5 = wb.create_sheet("标准配比参考")
ws5.merge_cells('A1:H1')
ws5.cell(row=1, column=1, value="标准工时配比体系 — 开发:测试:产品=6:3:2 | 前端:后端:架构=4:5:1").font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")

ws5.merge_cells('A3:H3')
ws5.cell(row=3, column=1, value="配比标准").font = Font(name="微软雅黑", size=11, bold=True)
ws5.cell(row=3, column=1).fill = sub_fill

ratio_data = [
    ["层级", "类型", "配比", "占比", "包含的任务类型", "", "", ""],
    ["一级配比", "开发", "6", "54.55%", "前端开发、后端开发、开发任务(架构)、技术设计", "", "", ""],
    ["", "测试", "3", "27.27%", "测试、测试-", "", "", ""],
    ["", "产品(需求)", "2", "18.18%", "需求设计、需求调研、需求验收、管理事务、运维、其他、UI设计", "", "", ""],
    ["", "", "", "", "", "", "", ""],
    ["二级配比\n(开发内)", "前端开发", "4", "40%", "占开发工时的40% (=总工时的21.8%)", "", "", ""],
    ["", "后端开发", "5", "50%", "占开发工时的50% (=总工时的27.3%)", "", "", ""],
    ["", "架构开发", "1", "10%", "占开发工时的10% (=总工时的5.5%)", "", "", ""],
    ["", "", "", "", "", "", "", ""],
    ["三级配比\n(产品内)", "纯产品", "85%", "—", "需求设计/调研/验收/管理事务", "", "", ""],
    ["", "UI设计", "15%", "—", "UI设计", "", "", ""],
]
for i, row_data in enumerate(ratio_data, 4):
    for j, val in enumerate(row_data):
        c = ws5.cell(row=i, column=j+1, value=val)
        c.font = n_font
        c.border = thin_border

# 参考值速查表
ws5.cell(row=17, column=1, value="不同总工时下的标准参考值速查").font = Font(name="微软雅黑", size=11, bold=True)
ws5.cell(row=17, column=1).fill = sub_fill
ws5.merge_cells('A17:H17')

ref_h = ["需求总工时", "开发(54.5%)", "测试(27.3%)", "产品(18.2%)", "前端(21.8%)", "后端(27.3%)", "架构(5.5%)", "UI(2.7%)"]
set_header(ws5, 18, ref_h)

for idx, total in enumerate([50, 100, 150, 200, 250, 300, 350, 400, 450], 19):
    std = calc_standard_ratios(min(total, THRESHOLD))
    vals = [total, std['开发'], std['测试'], std['产品'], std['前端开发'], std['后端开发'], std['架构开发'], std['UI']]
    set_row(ws5, idx, vals)

auto_w(ws5)

# ──────────────────────────────────────────────────────
# 保存
# ──────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"输出文件: {OUTPUT_PATH}")

# ════════════════════════════════════════════════════════
# 关键统计
# ════════════════════════════════════════════════════════
print("\n" + "="*60)
print("关键统计")
print("="*60)
print(f"超标需求: {len(over_reqs)} 个, 超标总工时: {over_reqs['超标量'].sum():.0f}h")
print(f"重分配方案: {len(plan)} 条, 转移总工时: {sum(a for _,_,a,_ in plan):.0f}h")
print(f"调整后仍超标: {still_over} 个")

# 匹配方式统计
print("\n匹配方式分布:")
for m in sorted(method_stats.keys(), key=lambda x: -method_stats[x]['total']):
    s = method_stats[m]
    print(f"  {m}: {s['count']}条, {s['total']:.0f}h")

# 仍超标的需求
if still_over > 0:
    print(f"\n仍超标需求:")
    remaining = req[req['调整后是否超标']].sort_values('调整后总工时', ascending=False)
    for _, row in remaining.iterrows():
        print(f"  {row['调整后总工时']:.0f}h (-{row['总工时']-row['调整后总工时']:.0f}h)  {row['相关研发需求'][:60]}")

# 开发占比偏离最大的需求
print("\n开发占比偏离最大的 TOP5:")
req['开发占比'] = (req['开发工时'] / req['总工时'] * 100).fillna(0)
top_dev = req[req['是否超标']].nlargest(5, '开发占比')
for _, row in top_dev.iterrows():
    print(f"  {row['相关研发需求'][:50]}... 开发占{row['开发占比']:.0f}%(目标54.5%), 需减{row['开发偏差']:.0f}h")
